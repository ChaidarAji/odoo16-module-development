import logging
import pytz
import math

from collections import namedtuple, defaultdict

from datetime import datetime, timedelta, time
from pytz import timezone, UTC
from odoo import models, fields, api, _
from odoo.exceptions import UserError

import openpyxl
import base64
from io import BytesIO

from dateutil.relativedelta import relativedelta
from odoo.tools import float_is_zero
from odoo.tools import date_utils
import io
import json

try:
   from odoo.tools.misc import xlsxwriter
except ImportError:
   import xlsxwriter


_logger = logging.getLogger(__name__)

class HrEmployeeContract(models.Model):
	_inherit = "hr.contract"
	
	acting_allowance 		= fields.Monetary(string='Acting Allowance', default= 0.00)
	competency_allowance 	= fields.Monetary(string='Competency Allowance', default= 0.00)
	location_allowance 		= fields.Monetary(string='Location Allowance', default= 0.00)
	position_allowance 		= fields.Monetary(string='Position Allowance', default= 0.00)
	daily_allowance 		= fields.Monetary(string='Daily Allowance', default= 0.00)
	handphone_allowance     = fields.Monetary(string='Handphone Allowance', default= 0.00)
	performance_allowance   = fields.Monetary(string='Performance Allowance', default= 0.00)
	housing_allowance   	= fields.Monetary(string='Housing Allowance', default= 0.00)
	shift_allowance   		= fields.Monetary(string='Long SHift Allowance', default= 0.00)
	tor_allowance   		= fields.Monetary(string='Tor Allowance', default= 0.00)
	guarantine_allowance   	= fields.Monetary(string='Guaranteed Incentive', default= 0.00)
	 
	# there will be other allowance and deduction here
	adjustment_plus			= fields.Monetary(string='Adjustment', default= 0.00)  
	all_backup				= fields.Monetary(string='All Backup', default= 0.00)  
	national_allowance		= fields.Monetary(string='Insentif Libur Nasional', default= 0.00)  
	netral_allowance		= fields.Monetary(string='Netral Allowance', default= 0.00)  
	rapel_allowance			= fields.Monetary(string='Rapel Allowance', default= 0.00)  
	tax_adjustment			= fields.Monetary(string='Tax Adjudtment', default= 0.00) 
	travel2_allowance		= fields.Monetary(string='Travel Allowance', default= 0.00)
	bonus					= fields.Monetary(string='Bonus', default= 0.00)
	car_allowance			= fields.Monetary(string='Car allowance', default= 0.00)

	# potongan
	deduct_adv				= fields.Monetary(string='Potongan Advance/Pot.Koperasi', default= 0.00)  
	deduct_car				= fields.Monetary(string='Potongan CAR', default= 0.00)  
	deduct_dormit			= fields.Monetary(string='Potongan Asrama', default= 0.00)  
	deduct_meal				= fields.Monetary(string='POTONGAN UANG MAKAN/POT MEDICAL', default= 0.00) 
	deduct_rapel			= fields.Monetary(string='Potongan Rapel', default= 0.00)  
	deduct_seragam			= fields.Monetary(string='Uang Jaminan Seragam', default= 0.00)  
	deduct_sepatu			= fields.Monetary(string='Uang Jaminan Sepatu', default= 0.00)  
	deduct_training			= fields.Monetary(string='Uang Jaminan Training', default= 0.00)  
	deduct_security			= fields.Monetary(string='Uang Jaminan Training Sec. Guard', default= 0.00)  



class HrPayrollDateSetting(models.Model):
	_name = 'hr.payroll.date.setting'
	# version 14
	name = fields.Char(string='Code')
	date = fields.Date(string='Date')
	

class HrPayrolRegionalSetting(models.Model):
	_name = 'hr.payroll.regional.setting'
	# version 14
	name 				= fields.Char(string='Name')
	description 		= fields.Char(string='Description')
	currency_id 		= fields.Many2one('res.currency', string="Currency",default=lambda self: self.env.user.company_id.currency_id.id)
	umr 				= fields.Monetary(string='Regional Value', default= 0.00)

class HrPayrolRegionalSetting(models.Model):
	_name = 'hr.payroll.regional.setting.value'
	
	name 				= fields.Many2one('hr.payroll.regional.setting', string="Regional")
	year 				= fields.Char(string='Tahun')
	currency_id 		= fields.Many2one('res.currency', string="Currency",default=lambda self: self.env.user.company_id.currency_id.id)
	value 				= fields.Monetary(string='Nilai', default= 0.00)
	state  = fields.Selection([('active', 'Active'), ('inactive', 'Inactive')], default="inactive")

	def button_active(self):
		for data in self:
			exist_umk_active = self.env['hr.payroll.regional.setting.value'].search([('name', '=', data.name.id), 
																			('id', '!=', data.id), 
																			('state', '=', 'active')])
			if exist_umk_active:
				exist_umk_active.write({'state': 'inactive'})
			
			data.state = 'active'


# transient model wizard
class ImportPayrollRegional(models.TransientModel):
	_name = "import.payroll.regional.wizard"
   
	file = fields.Binary(string="File", required=True)

	def import_regional_save(self):
		try :
			wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)

			ws = wb.active
			for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
				regional_info = self.env['hr.payroll.regional.setting'].sudo().search([('name','=',record[0])])
				#_logger.error('REGIONAL INFO')
				#_logger.error(regional_info)
				valid = True

				if len(regional_info) > 0:
					for regi in regional_info:
						# if exist update not exist create
						value_info = self.env['hr.payroll.regional.setting.value'].sudo().search([('name','=', regi.id),('year','=', record[1])])

						if value_info.id == False:
							res = self.env['hr.payroll.regional.setting.value'].sudo().create({
								'name' 	: regi.id,
								'year' 	: record[1],
								'value'	: record[2]
							})
						else:
							for valu in value_info:
								valu.write({
									'value' : record[2]
								})
				else:
					# insert dulu referensin
					res2 = self.env['hr.payroll.regional.setting'].sudo().create({
						'name'			: record[0],
						'description'	: record[0],
						'umr'			: 0
					})

					value_info = self.env['hr.payroll.regional.setting.value'].sudo().search([('name','=', res2.id),('year','=', record[1])])

					if value_info.id == False:
						res = self.env['hr.payroll.regional.setting.value'].sudo().create({
							'name' 	: res2.id,
							'year' 	: record[1],
							'value'	: record[2]
						})
					else:
						for valu in value_info:
							valu.write({
								'value' : record[2]
							})

					#raise UserError(record[0]+' tidak ada referensinya')
				
			return {
		  		'type': 'ir.actions.close_wizard_refresh_view'
	  		}
				
		except:
			raise UserError(_('Please insert a valid file'))

class ExportPayrollKoreksiGajiWizard(models.TransientModel):
	_name = "export.payroll.koreksi.gaji.wizard"

	only_template = fields.Selection([('data', 'Data'),
								   ('template', 'Template')], string='Opsi Export', default='template')
	export_rule_ids = fields.One2many("export.payroll.koreksi.gaji.wizard.rule", 'wizard_id', string="Salary Components")
	excel_file = fields.Binary(string='Download Report Excel',)
	period_id = fields.Many2one("hr.periode", string="Payroll Periode")

	
	def generate_excel_template(self):
		output 						= io.BytesIO()
		workbook 					= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 						= workbook.add_worksheet()

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})

		current_row = 1

		column_list = [
						'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
					 ]
		

		sheet.set_column(0, 0, 20)
		sheet.set_column(1, 1, 50)
		sheet.write(0, 0, 'NIK', header)
		sheet.write(0, 1, 'Nama', header)
		
		column = 2
		for rule in self.export_rule_ids:
			sheet.write(0, column, rule.rule_id.code, header)
			column+=1
		
		workbook.close()
		output.seek(0)
		file_base64 = base64.b64encode(output.read())
		output.close()
		self.excel_file = file_base64
		return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=export.payroll.koreksi.gaji.wizard&field=excel_file&download=true&id=%s&filename=%s' % (self.id, "Koreksi_Nilai_Gaji_Template.xlsx"),
            'target': 'self'
        }

	def generate_excel_data(self):
		output 						= io.BytesIO()
		workbook 					= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 						= workbook.add_worksheet()

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})


		column_list = [
						'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
					 ]
		

		sheet.set_column(0, 0, 20)
		sheet.set_column(1, 1, 50)
		sheet.write(0, 0, 'NIK', header)
		sheet.write(0, 1, 'Nama', header)
		header_names = {}
		column = 2
		for rule_line in self.export_rule_ids:
			sheet.write(0, column, rule_line.rule_id.code, header)
			header_names[rule_line.rule_id.code] = column
			column+=1

		all_data = {}
		payslip_line_ids = self.env['hr.payslip.line'].sudo().search([('slip_id.state','=','draft'), ('slip_id.payroll_periode','=', self.period_id.id)], order='nik asc')
		employee_ids = payslip_line_ids.mapped('employee_id')
		for employee in employee_ids.with_progress(msg="Generating Data"):
			rule_data = {}
			for header in header_names.keys():
				#Get rule data
				payslip_line_ids = payslip_line_ids.filtered(lambda p: p.employee_id.id == employee.id and p.salary_rule_id.code == header)
				print('===========================',header, employee.name, payslip_line_ids[0].amount if payslip_line_ids else 0)
				rule_data[header] = payslip_line_ids[0].amount if payslip_line_ids else 0
			all_data[employee.id] = rule_data
		print('=========================', all_data)
		current_row = 1
		for employee, data in all_data.items():
			employee_id = self.env['hr.employee'].sudo().browse(employee)
			sheet.write(current_row, 0, employee_id.nip, cell_left)
			sheet.write(current_row, 1, employee_id.name, cell_left)
			for header, value in data.items():
				sheet.write(current_row, header_names[header], value, cell_right_number)
			current_row+=1
		
		workbook.close()
		output.seek(0)
		file_base64 = base64.b64encode(output.read())
		output.close()
		self.excel_file = file_base64
		return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=export.payroll.koreksi.gaji.wizard&field=excel_file&download=true&id=%s&filename=%s' % (self.id, "Koreksi_Nilai_Gaji_Data.xlsx"),
            'target': 'self'
        }

	def button_export_data(self):
		if self.only_template == "template":
			return self.generate_excel_template()
		else:
			return self.generate_excel_data()

class ExportPayrollKoreksiGajiWizardRule(models.TransientModel):
	_name = "export.payroll.koreksi.gaji.wizard.rule"

	wizard_id = fields.Many2one("export.payroll.koreksi.gaji.wizard")
	rule_id = fields.Many2one("hr.salary.rule", string="Name", domain="[('can_override','=',True)]")

class ImportPayrollKoreksiGajiWizard(models.TransientModel):
	_name = "import.payroll.koreksi.gaji.wizard"
   
	file = fields.Binary(string="File", required=True)
	filename = fields.Char(string="Filename")
	period_id = fields.Many2one("hr.periode", string="Payroll Periode", required=True)
	total_employee_count = fields.Integer(string="Jumlah Pegawai Terhitung")
	total_employee_processed = fields.Integer(string="Jumlah Pegawai Terproses")
	error_ids = fields.One2many("import.payroll.koreksi.gaji.error.wizard", "wizard_id", string="Data Error")
	is_override = fields.Boolean(string='Override', default = True)
	ignore_formula = fields.Boolean(string='Ignore Formula', default = False)

	def import_koreksi_gaji_save(self):
		#try :
		wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)

		ws = wb.active
		#get column name
		column_name = {}
		for record in ws.iter_rows(min_row=1, max_row=1, min_col=3,max_col=None, values_only=True):
			for index, header_name in enumerate(record):
				column_name[str(index)] = header_name

		all_data = []
		for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
			all_data.append(record)

		updated_line = []
		employee_count = 0
		for record in self.web_progress_iter(all_data, msg="Reading Excel Data"):
			nik = record[0]
			employee_name = record[1]
			data_gaji = record[2:]
			for index, value in enumerate(data_gaji):
				payslip_line_info = self.env['hr.payslip.line'].sudo().search([('slip_id.state','=','draft'),('nik','=', nik),('code','=', column_name[str(index)]), ('slip_id.payroll_periode', '=', self.period_id.id)])
			#if len(payslip_line_info) > 0:
				if payslip_line_info:
					for line in payslip_line_info:
						line.write({
							'is_override' : self.is_override,
							'amount_correction' : float(value),
							'ignore_formula' : self.ignore_formula
						})
						updated_line.append(line.id)
				else:
					self.error_ids = [(0,0,{'nik' : nik, 'employee' : employee_name, 'name' : 'Komponen Gaji [%s] tidak ditemukan'%(column_name[str(index)])})]
			employee_count+=1

		self.total_employee_count = employee_count
		payslip_line_ids = self.env['hr.payslip.line'].sudo().browse(updated_line)
		employee_processed_count = 0
		employee_processed = []
		if payslip_line_ids:
			slip_id = payslip_line_ids.mapped('slip_id')
			for slip in slip_id.with_progress(msg="Processing Data Koreksi Nilai Gaji"):
				slip.compute_sheet_update()
				if slip.employee_id.id not in employee_processed:
					employee_processed_count += 1
					employee_processed.append(slip.employee_id.id)
		
		self.total_employee_processed = employee_processed_count

		if len(self.error_ids) > 0 :
			return { 
				'context'	: self.env.context, 
				'view_type'	: 'form', 
				'view_mode'	: 'form', 
				'res_model'	: 'import.payroll.koreksi.gaji.wizard', 
				'res_id'	: self.id, 
				'type'		: 'ir.actions.act_window', 
				'target'	: 'new' 
			}

		#except:
		#	raise UserError(_('Please insert a valid file'))

		#try :
		#	wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)

		#	ws = wb.active
		#	for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
				
				
				#regional_info = self.env['hr.payroll.regional.setting'].sudo().search([('name','=',record[0])])

				#valid = True

				#if len(regional_info) > 0:
				#	for regi in regional_info:
						# if exist update not exist create
				#		value_info = self.env['hr.payroll.regional.setting.value'].sudo().search([('name','=', regi.id),('year','=', record[1])])

				#		if value_info.id == False:
				#			res = self.env['hr.payroll.regional.setting.value'].sudo().create({
				#				'name' 	: regi.id,
				#				'year' 	: record[1],
				#				'value'	: record[2]
				#			})
				#		else:
				#			for valu in value_info:
				#				valu.write({
				#					'value' : record[2]
				#				})
				#else:
				#	raise UserError(record[0]+' tidak ada referensinya')
				
			#return {
		  	#	'type': 'ir.actions.close_wizard_refresh_view'
	  		#}
				
		#except:
		#	raise UserError(_('Please insert a valid file'))

class ImportPayrollKoreksiGajiErrorWizard(models.TransientModel):
	_name = "import.payroll.koreksi.gaji.error.wizard"

	nik = fields.Char(string="NIK")
	employee = fields.Char(string="Pegawai")
	name = fields.Text(string="Keterangan")
	wizard_id = fields.Many2one("import.payroll.koreksi.gaji.wizard")

class FilterPrePayrollWizard(models.TransientModel):
	_name = "filter.pre.payroll.wizard"

	year 	= fields.Integer(string='Tahun', Default  = int(datetime.now().strftime("%Y")))

	month    = fields.Selection(
		[ 
			('01', 'January'),
			('02', 'February'),
			('03', 'March'),
			('04', 'April'),
			('05', 'May'),
			('06', 'June'),
			('07', 'July'),
			('08', 'August'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'December')
		], 
		string   = 'Month',    
		required =False, 
		Default  = datetime.now().strftime("%m"))
	
	periode 		= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	name			= fields.Many2one(string = "Pegawai" ,comodel_name = "hr.employee")
	nik				= fields.Char(string='NIK')
	rule_id			= fields.Many2one(string = "Komponen Payroll" ,comodel_name = "hr.salary.rule", domain="[('can_override','=', True),('active','=', True)]")
	status			= fields.Selection([('active','Aktif'),('inactive','Tidak Aktif')], string='Status Payroll', default='active')

	def search_last(self):
		current_uid = self.env.uid

		wizard_info = self.env['filter.pre.payroll.wizard'].sudo().search_read([('create_uid','=',current_uid)], order="create_date desc", limit=1)

		data = {
			'status' : 200,
			'data'	 : wizard_info
		}


		return data
	
	def filter_pre_payroll_save(self):
		current_uid = self.env.uid
		search_info = self.env['ir.ui.view'].sudo().search([('name','=','internal_memo.hr_pre_payroll_view_search')])
		
		# updating filter first to match filtering
		#raise UserError(str(current_uid))
		wizard_info = self.env['filter.pre.payroll.wizard'].sudo().search([('create_uid','=',current_uid)], order="create_uid desc", limit=1)

		if wizard_info.id != False:
			wizard_info.write(
				{
					'year' 					: self.year,
					'month'					: self.month,
					'periode'				: self.periode.id,
					'name'					: self.name.id,
					'nik'					: self.nik,
					'rule_id'				: self.rule_id.id,
					'status'				: self.status
				}
			)
		else:
			self.env['filter.pre.payroll.wizard'].create({
				'year' 					: self.year,
				'month'					: self.month,
				'periode'				: self.periode.id,
				'name'					: self.name.id,
				'nik'					: self.nik,
				'rule_id'				: self.rule_id.id,
				'status'				: self.status
			})

		# name based on filter
		string_filter = ''

		domain = []

		if self.month != False:
			if self.month =='01':
				string_filter = string_filter + 'Januari'
			elif self.month =='02':
				string_filter = string_filter + 'Februari'
			elif self.month =='03':
				string_filter = string_filter + 'Maret'
			elif self.month =='04':
				string_filter = string_filter + 'April'
			elif self.month =='05':
				string_filter = string_filter + 'Mei'
			elif self.month =='06':
				string_filter = string_filter + 'Juni'
			elif self.month =='07':
				string_filter = string_filter + 'Juli'
			elif self.month =='08':
				string_filter = string_filter + 'Agustus'
			elif self.month =='09':
				string_filter = string_filter + 'September'
			elif self.month =='10':
				string_filter = string_filter + 'Oktober'
			elif self.month =='11':
				string_filter = string_filter + 'November'
			elif self.month =='12':
				string_filter = string_filter + 'Desember'

			domain.append(('bulan','=',self.month))

		

		if self.year != False:
			string_filter = string_filter + ' '+str(self.year)
			domain.append(('year','=',self.year))


		if self.periode.id != False:
			string_filter = string_filter + ' '+self.periode.name
			domain.append(('periode','=',self.periode.id))

		if self.name.id != False:
			string_filter = string_filter + ' '+ self.name.name
			domain.append(('name','=',self.name.id))

		if self.nik != False:
			string_filter = string_filter + ' '+ self.nik
			domain.append(('nik','=',self.nik))

		if self.rule_id.id != False:
			string_filter = string_filter + ' '+ self.rule_id.name
			domain.append(('rule_id','=',self.rule_id.id))

		if self.status != False:
			string_filter = string_filter + ' '+ self.status
			domain.append(('status','=',self.status))

		
		return {
			'type'					: 'ir.actions.act_window',
			'name'					: 'Payroll Data '+string_filter,
			'res_model'				: 'hr.pre.payroll',
			'view_id'				: False,
			'views'					: [(False, 'tree'),(False,'form')],
			'view_type' 			: 'tree',
			'view_mode' 			: 'tree,form',
			'target' 				: 'main',
			'nodestroy' 			: False,
			'search_view_id'      	: search_info.id,
			'domain'				: domain
			#'context'             	: { 'search_default_payroll_advanced': 1 }
		}
	


class ExportPrePayrollWizard(models.TransientModel):
	_name 			= "export.pre.payroll.wizard"
	_description    = "Export Data Process Payroll"

	name 			= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	only_template	= fields.Selection([('data','Dengan Data'),('template','Hanya Template')], string='Opsi Export')
	is_active		= fields.Selection([('active','Periode Aktif'),('other','Periode Lain')], string='Status')
	year 			= fields.Integer(string='Tahun', Default  = int(datetime.now().strftime("%Y")))

	month    = fields.Selection(
		[ 
			('01', 'January'),
			('02', 'February'),
			('03', 'March'),
			('04', 'April'),
			('05', 'May'),
			('06', 'June'),
			('07', 'July'),
			('08', 'August'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'December')
		], 
		string   = 'Bulan',    
		required = False, 
		Default  = datetime.now().strftime("%m"))
	
	rule_ids  = fields.One2many('export.pre.payroll.component.wizard', 'name', string='Salary Rule', readonly=False)


	# based onchange event
	@api.onchange('is_active')
	def onchange_is_active(self):
		# find based on date
		if self.name.id != False:
			periode_info 	= self.env['hr.periode'].sudo().search([('id','=', self.name.id)])
			self.year 		= periode_info.salaryenddate.year
			self.month		= str(periode_info.salaryenddate.month).zfill(2)


	def export_pre_payroll_save(self):
		if self.year == False:
			year 	= int(datetime.now().strftime("%Y"))
		else:
			year 	= self.year

		if self.year == False:
			month 	= datetime.now().strftime("%m")
		else:
			month 	= self.month

		if self.name == False:
			payroll_periode 	= False
		else:
			payroll_periode 	= self.name.id

		if self.only_template == False:
			only_template 	= 'template'
		else:
			only_template 	= self.only_template

		if self.is_active == False:
			is_active 	= 'active' # by default active periode
		else:
			is_active 	= self.is_active

		if len(self.rule_ids) <= 0:
			rule_ids 	= []
		else:
			rule_ids 	= []

			for ruleid in self.rule_ids:
				if ruleid.rule_id.id != False:
					rule_ids.append(ruleid.rule_id.id)

		data = {
		   	'year'					: year,
		   	'month'					: month,
		   	'payroll_periode'		: payroll_periode,
			'only_template'			: only_template,
			'is_active'				: is_active,
			'rule_ids'				: rule_ids
	   	}

		return {
		   'type'	: 'ir.actions.report',
		   'data'	: 
		   			{
						'model'				: 'export.pre.payroll.wizard',
						'options'			: json.dumps(data,default=date_utils.json_default),
						'output_format'		: 'xlsx',
						'report_name'		: 'Import Template Payroll Process',
					},
		   'report_type': 'xlsx',
	   }
	
	def get_pre_payroll_excel(self, data, response):
		# generate excel
		year 						= data['year']
		month 						= data['month']
		payroll_periode 			= data['payroll_periode']
		only_template 				= data['only_template']
		is_active 					= data['is_active']
		rule_ids 					= data['rule_ids']

		output 						= io.BytesIO()
		workbook 					= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 						= workbook.add_worksheet()

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})

		current_row = 1

		column_list = [
						'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
					 ]
		

		sheet.set_column(0, 0, 20)
		sheet.set_column(1, 1, 50)
		sheet.write('A'+str(current_row), 'NIK', header)
		sheet.write('B'+str(current_row), 'Nama', header)

		salary_rule = []

		currect_col 	= 2
		currect_col2    = 0

		for ruleid in rule_ids:
			# find salary rule
			rule_info = self.env['hr.salary.rule'].sudo().search([('id','=',ruleid)])

			sheet.set_column(currect_col, currect_col, 30)
			sheet.write(column_list[currect_col2]+str(current_row), rule_info.code, header)
			currect_col 	= currect_col + 1
			currect_col2 	= currect_col2 + 1


		current_row = current_row + 1

		# generate data on not only
		if only_template =='data':
			if is_active =='active':
				data_pre = self.env['hr.pre.payroll'].sudo().search([('periode','=',payroll_periode),('status','=','active'),('rule_id','in', rule_ids)])
			else:
				data_pre = self.env['hr.pre.payroll'].sudo().search([('periode','=',payroll_periode),('year','=',year),('bulan','=', month),('rule_id','in', rule_ids)])

			buffer_convert = {}

			if len(data_pre) > 0:
				#geser ke kanan, convert to array first
				for dat in data_pre:
					# jika object nggak ada create
					if 'item_'+str(dat.name.id) not in buffer_convert: 
						buffer_convert['item_'+str(dat.name.id)] = {}

					buffer_convert['item_'+str(dat.name.id)]['nik'] = dat.name.nip
					buffer_convert['item_'+str(dat.name.id)]['name'] = dat.name.name
					buffer_convert['item_'+str(dat.name.id)]['component_'+str(dat.rule_id.id)] = dat.value

			
			if len(buffer_convert) > 0:
				for dat2 in buffer_convert:
					currect_col2    = 0
					sheet.write('A'+str(current_row), buffer_convert[dat2]['nik'], cell_left)
					sheet.write('B'+str(current_row), buffer_convert[dat2]['name'], cell_left)

					for key in rule_ids:
						sheet.write(column_list[currect_col2]+str(current_row), buffer_convert[dat2]['component_'+str(key)], cell_right_number)
						currect_col2 = currect_col2 + 1

					current_row = current_row + 1



		workbook.close()
		output.seek(0)
		response.stream.write(output.read())
		output.close()
	


class ImportPrePayrollComponentWizard(models.TransientModel):
	_name 			= "export.pre.payroll.component.wizard"

	name 			= fields.Many2one(string = "Wizard ID" ,comodel_name = "export.pre.payroll.wizard") 
	rule_id			= fields.Many2one(string = "Komponen Payroll" ,comodel_name = "hr.salary.rule", domain="[('can_override','=',True)]")

		
class ImportPrePayrollWizard(models.TransientModel):
	_name 	= "import.pre.payroll.wizard"
   
	name 				= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode", required=True) 
	file 				= fields.Binary(string="File", required=True)
	ignore_formula		= fields.Boolean(string='Ignore Formula')
	all_employee_number	= fields.Integer(string='Jumlah Pegawai Terhitung')
	employee_number		= fields.Integer(string='Jumlah Pegawai Terproses')
	processed			= fields.Float(string='Progress (%)')

	error_ids 			= fields.One2many('import.pre.payroll.error.wizard', 'name', string='Data Error', readonly=False)
	counter_ids 		= fields.One2many('import.pre.payroll.counter.wizard', 'name', string='Data Error', readonly=True)

	def import_pre_payroll_save(self):
		# defining year and month
		periode_info 	= self.env['hr.periode'].sudo().search([('id','=', self.name.id)])

		wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)

		ws 				= wb.active
		counter 		= 0
		header_id		= {}
		counter_list 	= [(5,0,0)]
		record_saved	= []
		record_error	= [(5,0,0)]

		# extract year and month
		tahun	= periode_info.salaryenddate.year
		bulan	= str(periode_info.salaryenddate.month).zfill(2)


		# find all hr periode first to get data
		for record in ws.iter_rows(min_row=1, max_row=1, min_col=None,max_col=None, values_only=True):
			for x in range(2,20):
				try:
					if record[x] == '' or str(record[x]) =='None':
						continue

					salary_rule_info = self.env['hr.salary.rule'].sudo().search([('code','=',str(record[x]))])

					if salary_rule_info.id == False:
						raise UserError('Payroll component not found :'+str(record[x]))
					else:
						# simpan pada array
						header_id['header_'+str(x)] = salary_rule_info.id
				except IndexError:
					continue
		
		for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
			if record[0] == '' or str(record[0]) =='None':
				continue

			counter = counter + 1

			counter_list.append((0,0,{'counter' : counter,'name' : self.id}))
			record_saved.append(record)

			self.all_employee_number = counter

		self.write(
			{
				'counter_ids' : counter_list
			}
		)


		counter = 0
		for counte in self.counter_ids.with_progress(msg="Processing Payroll Data"):
			emp_info = self.env['hr.employee'].sudo().search([('nip','=', record_saved[counte.counter - 1][0])])
			
			if not emp_info:
				#raise UserError('Employee not found :'+str(record[0]))
				record_error.append((0,0,{'name' : self.id, 'nik' : record_saved[counte.counter - 1][0], 'employee' : record_saved[counte.counter - 1][1], 'description' : 'Tidak ditemukan'}))
				continue
				

			for x in range(2,20):
				try:
					if record[x] == '' or str(record[x]) =='None':
						continue

					# find active salary rule
					data_active = self.env['hr.pre.payroll'].sudo().search([('name','=', emp_info.id),('status','=','active'),('rule_id','=', header_id['header_'+str(x)])])

					if data_active.id == False:
						# create
						res = self.env['hr.pre.payroll'].sudo().create({
							'periode'		: self.name.id,
							'name' 			: emp_info.id,
							'rule_id'		: header_id['header_'+str(x)],
							'value'			: record[x],
							'status'		: 'active',
							'year'			: tahun,
							'bulan'			: bulan,
							'ignore_formula': self.ignore_formula,
							'nik'			: emp_info.nip
						})

						
					else:
						data_active.sudo().write({
							'periode'		: self.name.id,
							'name' 			: emp_info.id,
							'rule_id'		: header_id['header_'+str(x)],
							'value'			: record[x],
							'status'		: 'active',
							'year'			: tahun,
							'bulan'			: bulan,
							'ignore_formula': self.ignore_formula,
							'nik'			: emp_info.nip
						})

						
				except IndexError:
					continue

			counter 				= counter + 1
			self.employee_number  	= counter
			#self.error_ids			= record_error
			
		#raise UserError(str(len(record_error)))
		#_logger.error('CHECK LOG')
		#_logger.error(record_error)
		self.write(
			{
				'error_ids' : record_error
			}
		)


		# error will be written
		if len(record_error) > 0 :
			return { 
				'context'	: self.env.context, 
				'view_type'	: 'form', 
				'view_mode'	: 'form', 
				'res_model'	: 'import.pre.payroll.wizard', 
				'res_id'	: self.id, 
				'type'		: 'ir.actions.act_window', 
				'target'	: 'new' 
			}
		else:
			return { 
				'context'	: self.env.context, 
				'view_type'	: 'tree', 
				'view_mode'	: 'tree', 
				'res_model'	: 'hr.pre.payroll', 
				'type'		: 'ir.actions.act_window', 
				'target'	: 'new' 
			}		
	

class ImportPrePayrollErrorWizard(models.TransientModel):
	_name 			= "import.pre.payroll.error.wizard"

	name 			= fields.Many2one(string = "Wizard ID" ,comodel_name = "import.pre.payroll.wizard") 
	nik				= fields.Char(string='NIK')	
	employee		= fields.Char(string='Pegawai')
	description		= fields.Text(string='Keterangan')

class ImportPrePayrolCounterlWizard(models.TransientModel):
	_name 			= "import.pre.payroll.counter.wizard"

	name 			= fields.Many2one(string = "Wizard ID" ,comodel_name = "import.pre.payroll.wizard") 	
	counter			= fields.Integer(string='Counter')


class ImportContract(models.TransientModel):
	_name = "import.contract.wizard"
   
	file = fields.Binary(string="File", required=True)

	def import_contract_save(self):
		#try :
		wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)

		ws = wb.active
		for record in ws.iter_rows(min_row=2, max_row=None, min_col=None,max_col=None, values_only=True):
			if record[0] == '' or str(record[0]) =='None':
				continue

			# find employee
			emp_info = self.env['hr.employee'].sudo().search([('nip','=', record[0])])

			if not emp_info:
				raise UserError('Employee not found')
				
			contract_id  = self.env['hr.contract'].search([('id','=',emp_info.contract_id.id)])

			#_logger.error('EMPLOYEE INFO STATUS')
			#_logger.error(emp_info.employee_status_id.id)
	   		# struktur will be based on current employee status
			#raise UserError(emp_info.employee_status_id.id)
			periode_info = self.env['hr.periode'].sudo().search([('status_id','=', emp_info.employee_status_id.id),('category_id.code','=','DEFAULT')])

			is_periode_info = False
			#_logger.error('PERIODE INFO')
			#_logger.error(periode_info)

			for period in periode_info:
				is_periode_info = True

				# find default salary structure
				struct_id = self.env['hr.payroll.structure'].sudo().search([('periode_id','=',period.id)])
		
			if is_periode_info == False:
				raise UserError('Periode Payroll Belum Di Definisikan')
				
			# validasi semuanya
			if str(record[2]) =='None':
				raise UserError('Career Transition tidak valid ')
			
			if str(record[3]) =='None':
				raise UserError('Tipe Career Transition tidak valid ')

			# find transition
			trans_info = self.env['hr.employee.history.transition'].sudo().search([('name','=',record[2])])

			if trans_info.id == False:
				raise UserError('Career Transition tidak valid ')

			action_info	= self.env['hr.employee.history.action'].sudo().search([('name','=',record[3])])

			if action_info.id == False:
				raise UserError('Tipe Career Transition tidak valid ')

			if record[4] =='' or str(record[4]) =='None':
				employment_status_id = contract_id.contract_type_id.id
			else:
				# find nes one
				employee_status_info = self.env['hr.contract.type'].sudo().search([('name','=', record[4])])
				employment_status_id = employee_status_info.id

			# force to date
			new_contract_date = datetime.strptime(record[6], '%Y-%m-%d')

			if record[7] =='' or str(record[7]) =='None':
				new_contract_end = False
			else:
				new_contract_end = datetime.strptime(record[7], '%Y-%m-%d')

			

			if record[9] =='' or str(record[9]) =='None':
				new_master_id 	= contract_id.master_id.id
				new_job_id		= contract_id.job_id.id
			else:
				new_master_info = self.env['master.job'].sudo().search([('code','=',record[9])])

				if new_master_info:
					new_master_id = new_master_info.id
				else:
					new_master_info = self.env['master.job'].sudo().search([('name','=',record[9])])

					if new_master_info:
						new_master_id = new_master_info.id

				# job id
				new_job_info = self.env['hr.job'].sudo().search([('master_id','=',new_master_id)])
				new_job_id	 = new_job_info.id

				# diisi tapi salah
				if new_master_id == False:
					raise UserError('Job Position tidak valid')

			if record[10] =='' or str(record[10]) =='None':
				new_dept_id = contract_id.department_id.id
			else:
				new_dept_info 	= self.env['hr.department'].sudo().search([('name','=', record[10])])
				new_dept_id		= new_dept_info.id 

				if new_dept_id == False:
					raise UserError('Department tidak valid')


			# job status
			if record[11] =='' or str(record[11]) =='None':
				new_job_status_id = contract_id.job_status_id.id
			else:
				new_job_status_info = self.env['master.job.status'].sudo().search([('name','=', record[11])])
				new_job_status_id	= new_job_status_info.id

				if new_job_status_id == False:
					raise UserError('Job Status tidak valid')

			#work area
			if record[12] =='' or str(record[12]) =='None':
				new_area_id = contract_id.area_id.id
			else:
				# find
				new_area_info = self.env['area'].sudo().search([('code2','=',record[12])])

				if new_area_info:
					new_area_id = new_area_info.id
				else:
					new_area_info = self.env['area'].sudo().search([('name','=',record[12])])

					if new_master_info:
						new_area_id = new_area_info.id
					else:
						raise UserError('Cost Center tidak valid')

			# work location
			if record[13] =='' or str(record[13]) =='None':
				new_work_id = contract_id.work_location_id.id
				new_address_id = contract_id.work_location_id.address_id.id
			else:
				# find
				new_work_info = self.env['hr.work.location'].sudo().search([('code2','=',record[13])])

				if new_work_info:
					new_work_id = new_work_info.id
					new_address_id = new_work_info.address_id.id
				else:
					new_work_info = self.env['hr.work.location'].sudo().search([('name','=',record[13])])

					if new_work_info:
						new_work_id 	= new_work_info.id
						new_address_id 	= new_work_info.address_id.id
					else:
						raise UserError('Work Location tidak valid')

			# HR responsible
			if record[14] =='' or str(record[14]) =='None':
				new_responsible_id 	= contract_id.hr_responsible_id.id
			else:
				new_responsible_info = self.env['hr.employee'].sudo().search([('nip','=',record[14])])
				new_responsible_id	 = new_responsible_info.id

			if record[15] =='' or str(record[15]) =='None':
				new_grade_id 	= contract_id.grade_id.id
			else:
				new_grade_info = self.env['hr.grade'].sudo().search([('code','=',record[15])])

				if not new_grade_info:
					new_grade_info = self.env['hr.grade'].sudo().search([('name','=',record[15])])

				new_grade_id	 = new_grade_info.id

				if new_grade_id == False:
					raise UserError('Grade tidak valid')

			if record[16] =='' or str(record[16]) =='None':
				new_resign_id 	= False
			else:
				new_resign_info = self.env['hr.resignation.type'].sudo().search([('name','=',record[16])])
				new_resign_id	= new_resign_info.id

				if new_resign_id == False:
					raise UserError('Resign Type tidak valid '+record[16])

			if record[17] =='' or str(record[17]) =='None':
				new_resign_reason_id 	= False
			else:
				new_resign_info = self.env['hr.resignation.reason'].sudo().search([('name','=',record[17])])
				new_resign_reason_id	= new_resign_info.id

				if new_resign_reason_id == False:
					raise UserError('Resign Reason tidak valid')

			if record[18] =='' or str(record[18]) =='None':
				new_resign_date 	= False
			else:
				new_resign_date = datetime.strptime(record[18], '%Y-%m-%d')

			
			if record[19] =='' or str(record[19]) =='None':
				new_supervisor_id 	= contract_id.supervisor_id.id
			else:
				new_upervisor_info = self.env['hr.employee'].sudo().search([('nip','=',record[19])])
				new_supervisor_id	 = new_upervisor_info.id

				if new_supervisor_id == False:
					raise UserError('Supervisor tidak valid')

			if record[20] =='' or str(record[20]) =='None':
				new_manager_id 	= contract_id.manager_id.id
			else:
				new_manager_info = self.env['hr.employee'].sudo().search([('nip','=',record[20])])
				new_manager_id	 = new_manager_info.id

				if new_manager_id == False:
					raise UserError('Manager tidak valid')

			if record[22] =='' or str(record[22]) =='None':
				vacant_id = False
			else:
				if record[22] =='ya':
					vacant_id = True
				else:
					vacant_id = False

			if record[23] =='' or str(record[23]) =='None':
				salary_up_id = False
			else:
				if record[23] =='ya':
					salary_up_id = True
				else:
					salary_up_id = False

			

			if record[24] =='' or str(record[24]) =='None':
				gaji_pokok = contract_id.wage
			else:
				gaji_pokok	= record[24]

			# acting
			if record[25] =='' or str(record[25]) =='None':
				acting_allowance = contract_id.acting_allowance
			else:
				acting_allowance	= record[25]

			# competency
			if record[26] =='' or str(record[26]) =='None':
				competency_allowance = contract_id.competency_allowance
			else:
				competency_allowance	= record[26]

			if record[27] =='' or str(record[27]) =='None':
				location_allowance = contract_id.location_allowance
			else:
				location_allowance	= record[27]

			if record[28] =='' or str(record[28]) =='None':
				location_tsm_allowance = contract_id.allow_locationsm
			else:
				location_tsm_allowance	= record[28]

			# position
			if record[29] =='' or str(record[29]) =='None':
				position_allowance = contract_id.position_allowance
			else:
				position_allowance	= record[29]

			# daily allowance
			if record[30] =='' or str(record[30]) =='None':
				daily_allowance = contract_id.daily_allowance
			else:
				daily_allowance	= record[30]

			# tor allowance
			if record[31] =='' or str(record[31]) =='None':
				tor_allowance = contract_id.tor_allowance
			else:
				tor_allowance	= record[31]

			# handphone allowance
			if record[32] =='' or str(record[32]) =='None':
				handphone_allowance = contract_id.handphone_allowance
			else:
				handphone_allowance	= record[32]

			# performance allowance
			if record[33] =='' or str(record[33]) =='None':
				performance_allowance = contract_id.performance_allowance
			else:
				performance_allowance	= record[33]

			# transport allowance
			if record[34] =='' or str(record[34]) =='None':
				transport_allowance = contract_id.travel_allowance
			else:
				transport_allowance	= record[34]
			
			# meal allowance
			if record[35] =='' or str(record[35]) =='None':
				meal_allowance = contract_id.meal_allowance
			else:
				meal_allowance	= record[35]

			# medical allowance
			if record[35] =='' or str(record[35]) =='None':
				medical_allowance = contract_id.medical_allowance
			else:
				medical_allowance	= record[35]

			# hosign allowance
			if record[37] =='' or str(record[37]) =='None':
				housing_allowance = contract_id.housing_allowance
			else:
				housing_allowance	= record[37]

			# shift allowance
			if record[38] =='' or str(record[38]) =='None':
				shift_allowance = contract_id.shift_allowance
			else:
				shift_allowance	= record[38]

			# guarantine allowance
			if record[39] =='' or str(record[39]) =='None':
				guarantine_allowance = contract_id.guarantine_allowance
			else:
				guarantine_allowance	= record[39]


			# incentive allowance
			if record[40] =='' or str(record[40]) =='None':
				incentive_allowance = contract_id.allow_incentive
			else:
				incentive_allowance	= record[40]

			# adjustment allowance
			if record[41] =='' or str(record[41]) =='None':
				adjustment_allowance = contract_id.adjustment_plus
			else:
				adjustment_allowance	= record[41]

			# all backup allowance
			if record[42] =='' or str(record[42]) =='None':
				backup_allowance = contract_id.all_backup
			else:
				backup_allowance	= record[42]


			# national allowance
			if record[43] =='' or str(record[43]) =='None':
				national_allowance = contract_id.national_allowance
			else:
				national_allowance	= record[43]

			# netral allowance
			if record[44] =='' or str(record[44]) =='None':
				netral_allowance = contract_id.netral_allowance
			else:
				netral_allowance	= record[44]

			# rapel allowance
			if record[45] =='' or str(record[45]) =='None':
				rapel_allowance = contract_id.rapel_allowance
			else:
				rapel_allowance	= record[45]

			# tax adjustment allowance
			if record[46] =='' or str(record[46]) =='None':
				tax_adjust_allowance = contract_id.tax_adjustment
			else:
				tax_adjust_allowance	= record[46]

			# travel allowance
			if record[47] =='' or str(record[47]) =='None':
				travel_allowance = contract_id.travel2_allowance
			else:
				travel_allowance	= record[47]

			# bonus allowance
			if record[48] =='' or str(record[48]) =='None':
				bonus = contract_id.bonus
			else:
				bonus	= record[48]

			# car allowance
			if record[49] =='' or str(record[49]) =='None':
				car_allowance = contract_id.car_allowance
			else:
				car_allowance	= record[49]

			if record[50] =='' or str(record[50]) =='None':
				deduct_adv = contract_id.deduct_adv
			else:
				deduct_adv	= record[50]

			if record[51] =='' or str(record[51]) =='None':
				deduct_car = contract_id.deduct_car
			else:
				deduct_car	= record[51] 

			if record[52] =='' or str(record[52]) =='None':
				deduct_dormit = contract_id.deduct_dormit
			else:
				deduct_dormit	= record[52] 

			if record[53] =='' or str(record[53]) =='None':
				deduct_meal = contract_id.deduct_meal
			else:
				deduct_meal	= record[53] 

			if record[54] =='' or str(record[54]) =='None':
				deduct_rapel = contract_id.deduct_rapel
			else:
				deduct_rapel	= record[54] 

			if record[55] =='' or str(record[55]) =='None':
				deduct_seragam = contract_id.deduct_seragam
			else:
				deduct_seragam	= record[55] 

			if record[56] =='' or str(record[56]) =='None':
				deduct_sepatu = contract_id.deduct_sepatu
			else:
				deduct_sepatu	= record[56] 

			if record[57] =='' or str(record[57]) =='None':
				deduct_training = contract_id.deduct_training
			else:
				deduct_training	= record[57] 

			if record[58] =='' or str(record[58]) =='None':
				deduct_security = contract_id.deduct_security
			else:
				deduct_security	= record[58] 

			if record[59] =='' or str(record[59]) =='None':
				other_allowance = contract_id.other_allowance
			else:
				other_allowance	= record[59] 

			if record[60] =='' or str(record[60]) =='None':
				bpjs_ks_mitra = contract_id.bpjs_ks_mitra
			else:
				bpjs_ks_mitra	= record[60] 

			if record[61] =='' or str(record[61]) =='None':
				bpjs_tk_mitra = contract_id.bpjs_tk_mitra
			else:
				bpjs_tk_mitra	= record[61] 

			
			today 		= fields.Datetime.now()
			today 		= today + timedelta(hours=7)
			today_date 	= today.date()

			#raise UserError(str(today_date)+' '+str(new_contract_date))

			if today_date >= new_contract_date.date() and (new_contract_end == False or (new_contract_end != False and new_contract_end.date() >= today_date)):
				before_contract = self.env['hr.contract'].sudo().search([('state','=','open'),('employee_id','=',emp_info.id )])

				close_date = new_contract_date
				close_date = close_date - timedelta(days=1)

				#_logger.error('before contract')
				#_logger.error(before_contract)
				for befo in before_contract:
					befo.write({
						'state'			: 'close',
						'date_end'		: close_date
					})

			#raise UserError('SAMPAI DISINI')
			#raise UserError(str(contract_id.employee_status2_id.id))
		
			resx = self.env['hr.contract'].sudo().create({
				'employee_id'               : emp_info.id,
				'employee_id_updt'          : contract_id.employee_id.id,
				'date_start'                : new_contract_date,
				'date_start_updt'           : contract_id.date_start,
				'date_end'                  : new_contract_end,
				'date_end_updt'             : contract_id.date_end,
				'department_id'             : contract_id.department_id.id,
				'department_id_updt'        : new_dept_id,
				'master_id'                 : contract_id.master_id.id,
				'master_id_updt'            : new_master_id,
				#'notice_days'               : contract_id.notice_days,
				#'notice_days_updt'          : 0,
				'structure_type_id'         : 1, #contract_id.structure_type_id.id,
				'structure_type_id_updt'    : contract_id.structure_type_id.id,
				'resource_calendar_id'      : 1,
				'resource_calendar_id_updt' : contract_id.resource_calendar_id.id,
				#'tax_type_id'               : False,
				#'tax_type_id_updt'          : contract_id.tax_type_id.id,
				#'tax_location_id'           : False,
				#'tax_location_id_updt'      : contract_id.tax_location_id.id,
				'schedule_pay'              : 'monthly',
				'schedule_pay_updt'         : contract_id.schedule_pay,
				'struct_id'                 : struct_id.id,
				'struct_id_updt'            : contract_id.struct_id.id,
				'type_id'                   : employment_status_id,
				'type_id_updt'              : contract_id.type_id.id,
				'job_id'                    : new_job_id,
				'job_id_updt'               : contract_id.job_id.id,
				'contract_type_id'          : employment_status_id,
				'contract_type_id_updt'     : contract_id.contract_type_id.id,
				'hr_responsible_id'         : new_responsible_id,
				'hr_responsible_id_updt'    : contract_id.hr_responsible_id.id,
				'employee_status2_id'        : emp_info.employee_status_id.id,
				'employee_status_id2_updt'   : contract_id.employee_id.employee_status_id.id,
				'work_location_id'          : new_work_id,
				'work_location_id_updt'     : contract_id.work_location_id.id,
				'work_address_id'           : new_address_id,
				'work_address_id_updt'      : contract_id.work_address_id.id,
				'wage'                      : gaji_pokok,
				'wage_updt'                 : contract_id.wage,
				'tunjangan_jabatan'         : 0,
				'tunjangan_jabatan_updt'    : contract_id.tunjangan_jabatan,
				'acting_allowance'          : acting_allowance,
				'acting_allowance_updt'     : contract_id.acting_allowance,
				'competency_allowance'      : competency_allowance,
				'competency_allowance_updt' : contract_id.competency_allowance,
				'location_allowance'        : location_allowance,
				'location_allowance_updt'   : contract_id.location_allowance,
				'allow_locationsm'          : location_tsm_allowance,
				'allow_locationsm_updt'     : contract_id.allow_locationsm,
				'position_allowance'        : position_allowance,
				'position_allowance_updt'   : contract_id.position_allowance,
				'daily_allowance'           : daily_allowance,
				'daily_allowance_updt'      : contract_id.daily_allowance,
				'tor_allowance'             : tor_allowance,
				'tor_allowance_updt'        : contract_id.tor_allowance,
				'handphone_allowance'       : handphone_allowance,
				'handphone_allowance_updt'  : contract_id.handphone_allowance,
				'performance_allowance'     : performance_allowance,
				'performance_allowance_updt': contract_id.performance_allowance,
				'travel_allowance'          : transport_allowance,
				'travel_allowance_updt'     : contract_id.travel_allowance,
				'meal_allowance'            : meal_allowance,
				'meal_allowance_updt'       : contract_id.meal_allowance,
				'medical_allowance'         : medical_allowance,
				'medical_allowance_updt'    : contract_id.medical_allowance,
				'housing_allowance'         : housing_allowance,
				'housing_allowance_updt'    : contract_id.housing_allowance,
				'shift_allowance'           : shift_allowance,
				'shift_allowance_updt'      : contract_id.shift_allowance,
				'guarantine_allowance'      : guarantine_allowance,
				'guarantine_allowance_updt' : contract_id.guarantine_allowance,
				'allow_incentive'           : incentive_allowance,
				'allow_incentive_updt'      : contract_id.allow_incentive,
				'adjustment_plus'           : adjustment_allowance,
				'adjustment_plus_updt'      : contract_id.adjustment_plus,
				'all_backup'                : backup_allowance,
				'all_backup_updt'           : contract_id.all_backup,
				'national_allowance'        : national_allowance,
				'national_allowance_updt'   : contract_id.national_allowance,
				'netral_allowance'          : netral_allowance,
				'netral_allowance_updt'     : contract_id.netral_allowance,
				'rapel_allowance'           : rapel_allowance,
				'rapel_allowance_updt'      : contract_id.rapel_allowance,
				'tax_adjustment'            : tax_adjust_allowance,
				'tax_adjustment_updt'       : contract_id.tax_adjustment,
				'travel2_allowance'         : travel_allowance,
				'travel2_allowance_updt'    : contract_id.travel2_allowance,
				'bonus'                     : bonus,
				'bonus_updt'                : contract_id.bonus,
				'car_allowance'             : car_allowance,
				'car_allowance_updt'        : contract_id.car_allowance,
				'deduct_adv'                : deduct_adv,
				'deduct_adv_updt'           : contract_id.deduct_adv,
				'deduct_car'                : deduct_car,
				'deduct_car_updt'           : contract_id.deduct_car,
				'deduct_dormit'             : deduct_dormit,
				'deduct_dormit_updt'        : contract_id.deduct_dormit,
				'deduct_meal'               : deduct_meal,
				'deduct_meal_updt'          : contract_id.deduct_meal,
				'deduct_rapel'              : deduct_rapel,
				'deduct_rapel_updt'         : contract_id.deduct_rapel,
				'deduct_seragam'            : deduct_seragam,
				'deduct_seragam_updt'       : contract_id.deduct_seragam,
				'deduct_sepatu'             : deduct_sepatu,
				'deduct_sepatu_updt'        : contract_id.deduct_sepatu,
				'deduct_training'           : deduct_training,
				'deduct_training_updt'      : contract_id.deduct_training,
				'deduct_security'           : deduct_security,
				'deduct_security_updt'      : contract_id.deduct_security,
				'bpjs_ks_mitra'             : bpjs_ks_mitra,
				'bpjs_ks_mitra_updt'        : contract_id.bpjs_ks_mitra,
				'bpjs_tk_mitra'             : bpjs_tk_mitra,
				'bpjs_tk_mitra_updt'        : contract_id.bpjs_tk_mitra,
				'other_allowance'           : other_allowance,
				'other_allowance_updt'      : contract_id.other_allowance,
				'trans_id'             		: trans_info.id,
				'trans_id_updt'             : contract_id.trans_id.id,
				'action_id'            		: action_info.id,
				'action_id_updt'            : contract_id.action_id.id,
				'employee_status_id2_updt'   : contract_id.employee_status2_id.id,
				'grade_id_updt'             : contract_id.grade_id.id,
				'grade_id'                  : new_grade_id,
				'job_status_id_updt'        : contract_id.job_status_id.id,
				'job_status_id'             : new_job_status_id,
				'resign_type_id'       		: new_resign_id,
				'resign_type_id_updt'       : contract_id.resign_type_id.id,
				'resign_reason_id'     		: new_resign_reason_id,
				'resign_reason_id_updt'     : contract_id.resign_reason_id.id,
				'resign_date'	          	: new_resign_date,
				'resign_date_updt'          : contract_id.resign_date,
				'supervisor_id'        		: new_supervisor_id,
				'supervisor_id_updt'        : contract_id.supervisor_id.id,
				'manager_id'           		: new_manager_id,
				'manager_id_updt'           : contract_id.manager_id.id,
				'remark'               		: record[21],
				'remark_updt'               : contract_id.remark,
				'vacant_updt'               : contract_id.vacant,
				'area_id_updt'              : contract_id.area_id.id,
				'area_id'                   : new_area_id,
				'nik_updt'                  : contract_id.nik,
				'nik'                       : record[8],
				'request_no'           		: record[5],
				'request_no_updt'           : contract_id.request_no_updt,
				'state'						: 'draft'
			})

			

			if resx:
				# write, hanya jika periode aktif, jika peride kurang expired, jika periode lebih draft
				today 		= fields.Datetime.now()
				today 		= today + timedelta(hours=7)
				today_date 	= today.date()

				if today_date >= new_contract_date.date() and (new_contract_end == False or (new_contract_end != False and new_contract_end.date() >= today_date)):
					# on going
					resx.write({
						'state' : 'open'
					})

					# update employee
					emp_info.write({
						'nip' 				: record[8],
						'employment_status'	: employment_status_id,
						'job_status'		: new_job_status_id,
						'job_id'			: new_job_id,
						'master_id'			: new_master_id,
						'department_id'		: new_dept_id,
						'area'				: new_area_id,
						'work_location_id'	: new_work_id,
						'address_id'		: new_address_id
					})
				elif new_contract_end != False and new_contract_end.date() < today_date:
					resx.write({
						'state' : 'close'
					})
			

	


		# return here
		return {
		  	'type': 'ir.actions.close_wizard_refresh_view'
	  	}

		#except:
		#	raise UserError(_('Please insert a valid file'))


class FilterPayslip(models.TransientModel):
	_name = "filter.payslip.wizard"

	year 	= fields.Integer(string='Tahun', Default  = int(datetime.now().strftime("%Y")))

	month    = fields.Selection(
		[ 
			('01', 'January'),
			('02', 'February'),
			('03', 'March'),
			('04', 'April'),
			('05', 'May'),
			('06', 'June'),
			('07', 'July'),
			('08', 'August'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'December')
		], 
		string   = 'Month',    
		required =True, 
		Default  = datetime.now().strftime("%m"))
	
	category 			= fields.Many2one('hr.periode.category',string="Kategori Gaji", required=True)
	payroll_periode		= fields.Many2one('hr.employee.status',string="Payroll Periode")
	cost_center			= fields.Many2one('area',string="Cost Center")
	work_location		= fields.Many2one('hr.work.location',string="Work Location")
	pay_freq			= fields.Selection([('DAILY','DAILY'),('MONTHLY','MONTHLY')], string='Pay Freq')
	tax_location		= fields.Many2one('tax.location',string="Tax Location")
	tax_type			= fields.Selection([('local','Local'),('fixed','Fixed')], string='Tax Type')


	def search_last(self):
		current_uid = self.env.uid

		wizard_info = self.env['filter.payslip.wizard'].sudo().search_read([('create_uid','=',current_uid)], order="create_date desc", limit=1)

		data = {
			'status' : 200,
			'data'	 : wizard_info
		}


		return data

	def filter_payslip_save(self):
		current_uid = self.env.uid
		search_info = self.env['ir.ui.view'].sudo().search([('name','=','hr.payslip.custom2.model.view.search')])
		
		# updating filter first to match filtering
		#raise UserError(str(current_uid))
		wizard_info = self.env['filter.payslip.wizard'].sudo().search([('create_uid','=',current_uid)], order="create_uid desc", limit=1)

		if wizard_info.id != False:
			wizard_info.write(
				{
					'year' 					: self.year,
					'month'					: self.month,
					'category'				: self.category.id,
					'payroll_periode'		: self.payroll_periode.id,
					'cost_center'			: self.cost_center.id,
					'work_location'			: self.work_location.id,
					'pay_freq'				: self.pay_freq,
					'tax_location'			: self.tax_location.id,
					'tax_type'				: self.tax_type
				}
			)
		else:
			self.env['filter.payslip.wizard'].create({
				'year' 					: self.year,
				'month'					: self.month,
				'category'				: self.category.id,
				'payroll_periode'		: self.payroll_periode.id,
				'cost_center'			: self.cost_center.id,
				'work_location'			: self.work_location.id,
				'pay_freq'				: self.pay_freq,
				'tax_location'			: self.tax_location.id,
				'tax_type'				: self.tax_type
			})

		# name based on filter
		string_filter = ''

		domain = []

		if self.month != False:
			if self.month =='01':
				string_filter = string_filter + 'Januari'
			elif self.month =='02':
				string_filter = string_filter + 'Februari'
			elif self.month =='03':
				string_filter = string_filter + 'Maret'
			elif self.month =='04':
				string_filter = string_filter + 'April'
			elif self.month =='05':
				string_filter = string_filter + 'Mei'
			elif self.month =='06':
				string_filter = string_filter + 'Juni'
			elif self.month =='07':
				string_filter = string_filter + 'Juli'
			elif self.month =='08':
				string_filter = string_filter + 'Agustus'
			elif self.month =='09':
				string_filter = string_filter + 'September'
			elif self.month =='10':
				string_filter = string_filter + 'Oktober'
			elif self.month =='11':
				string_filter = string_filter + 'November'
			elif self.month =='12':
				string_filter = string_filter + 'Desember'

			domain.append(('month','=',self.month))

		

		if self.year != False:
			string_filter = string_filter + ' '+str(self.year)
			domain.append(('year','=',self.year))

		if self.category.id != False:
			string_filter = string_filter + ' '+self.category.name
			domain.append(('category','=',self.category.id))

		if self.payroll_periode.id != False:
			string_filter = string_filter + ' '+ self.payroll_periode.name
			domain.append(('payroll_periode','=',self.payroll_periode.id))

		if self.cost_center.id != False:
			string_filter = string_filter + ' '+ self.cost_center.name
			domain.append(('area_id','=',self.cost_center.id))

		if self.work_location.id != False:
			string_filter = string_filter + ' '+ self.work_location.name
			domain.append(('location_id','=',self.work_location.id))

		if self.pay_freq != False:
			string_filter = string_filter + ' '+ self.pay_freq
			domain.append(('payfreq','=',self.pay_freq))

		if self.tax_location.id != False:
			string_filter = string_filter + ' '+ self.tax_location.name
			domain.append(('tax_location_id','=',self.tax_location.id))

		if self.tax_type != False:
			string_filter = string_filter + ' '+ self.tax_type
			domain.append(('tax_type','=',self.tax_type))
		
		


		
		return {
			'type'					: 'ir.actions.act_window',
			'name'					: 'Payslip Data '+string_filter,
			'res_model'				: 'hr.payslip',
			'view_id'				: False,
			'views'					: [(False, 'tree'),(False,'form')],
			'view_type' 			: 'tree',
			'view_mode' 			: 'tree,kanban,form',
			'target' 				: 'main',
			'nodestroy' 			: False,
			'search_view_id'      	: search_info.id,
			'domain'				: domain
			#'context'             	: { 'search_default_payroll_advanced': 1 }
		}


class ExportPayslipExcel(models.TransientModel):
	_name = "export.payslip.wizard"

	# basen on chosen
	year 	= fields.Integer(string='Tahun', Default  = int(datetime.now().strftime("%Y")))

	month    = fields.Selection(
		[ 
			('01', 'January'),
			('02', 'February'),
			('03', 'March'),
			('04', 'April'),
			('05', 'May'),
			('06', 'June'),
			('07', 'July'),
			('08', 'August'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'December')
		], 
		string   = 'Month',    
		required =True, 
		Default  = datetime.now().strftime("%m"))
	
	category 			= fields.Many2one('hr.periode.category',string="Kategori Gaji", required=True)
	payroll_periode		= fields.Many2one('hr.employee.status',string="Payroll Periode")
	cost_center			= fields.Many2one('area',string="Cost Center")
	work_location		= fields.Many2one('hr.work.location',string="Work Location")
	pay_freq			= fields.Selection([('DAILY','DAILY'),('MONTHLY','MONTHLY')], string='Pay Freq')
	tax_location		= fields.Many2one('tax.location',string="Tax Location")
	tax_type			= fields.Selection([('local','Local'),('fixed','Fixed')], string='Tax Type')


	def export_payslip_save(self):
		if self.year == False:
			year 	= int(datetime.now().strftime("%Y"))
		else:
			year 	= self.year

		if self.year == False:
			month 	= datetime.now().strftime("%m")
		else:
			month 	= self.month

		if self.category == False:
			category 	= False
		else:
			category 	= self.category.id

		if self.payroll_periode.id == False:
			payroll_periode 	= False
		else:
			payroll_periode 	= self.payroll_periode.id

		if self.cost_center.id == False:
			cost_center 	= False
		else:
			cost_center 	= self.cost_center.id

		if self.work_location.id == False:
			work_location 	= False
		else:
			work_location 	= self.work_location.id

		if self.pay_freq == False:
			pay_freq 	= False
		else:
			pay_freq 	= self.pay_freq

		if self.tax_location.id == False:
			tax_location 	= False
		else:
			tax_location 	= self.tax_location.id

		if self.tax_type == False:
			tax_type 	= False
		else:
			tax_type 	= self.tax_type


		data = {
		   'year'					: year,
		   'month'					: month,
		   'category'				: category,
			'payroll_periode'		: payroll_periode,
			'cost_center'			: cost_center,
			'work_location'			: work_location,
			'pay_freq'				: pay_freq,
			'tax_location'			: tax_location,
			'tax_type'				: tax_type
	   	}

		return {
		   'type'	: 'ir.actions.report',
		   'data'	: 
		   			{
						'model'				: 'export.payslip.wizard',
						'options'			: json.dumps(data,default=date_utils.json_default),
						'output_format'		: 'xlsx',
						'report_name'		: 'Laporan Payroll',
					},
		   'report_type': 'xlsx',
	   }

	# using HTTP controller instead of default class ?
	def get_payslip_excel(self, data, response):
		year 					= data['year']
		month 					= data['month']
		category 				= data['category']
		payroll_periode 		= data['payroll_periode']
		cost_center 			= data['cost_center']
		work_location 			= data['work_location']
		pay_freq 				= data['pay_freq']
		tax_location 			= data['tax_location']
		tax_type 				= data['tax_type']

		output 					= io.BytesIO()
		workbook 				= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 					= workbook.add_worksheet()

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})
		
		category_info 			= self.env['hr.periode.category'].sudo().search([('id','=', int(category))])
		
		if month =='01':
			nama_bulan = 'JANUARI'
		elif month =='02':
			nama_bulan = 'FEBRUARI'
		elif month =='03':
			nama_bulan = 'MARET'
		elif month =='04':
			nama_bulan = 'APRIL'
		elif month =='05':
			nama_bulan = 'MEI'
		elif month =='06':
			nama_bulan = 'JUNI'
		elif month =='07':
			nama_bulan = 'JULI'
		elif month =='08':
			nama_bulan = 'AGUSTUS'
		elif month =='09':
			nama_bulan = 'SEPTEMBER'
		elif month =='10':
			nama_bulan = 'OKTOBER'
		elif month =='11':
			nama_bulan = 'NOVEMBER'
		elif month =='12':
			nama_bulan = 'DESEMBER'

		sheet.merge_range('A1:I1','LAPORAN PENGGAJIAN TAHUN '+str(year)+' BULAN '+nama_bulan, head)
		
		current_row = 3

		sheet.set_column(0, 0, 10)
		sheet.set_column(1, 1, 20)
		sheet.set_column(2, 2, 30)
		sheet.set_column(3, 3, 30)
		sheet.set_column(4, 4, 30)
		sheet.set_column(5, 5, 15)
		sheet.set_column(6, 6, 15)
		sheet.set_column(7, 7, 10)
		sheet.set_column(8, 8, 30)

		column_list = [
						'J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
					 ]

		sheet.write('A'+str(current_row), 'No', header)
		sheet.write('B'+str(current_row), 'NIK', header)
		sheet.write('C'+str(current_row), 'Nama', header)
		sheet.write('D'+str(current_row), 'Job Position', header)
		sheet.write('E'+str(current_row), 'Cost Center', header)
		sheet.write('F'+str(current_row), 'Payfreq', header)
		sheet.write('G'+str(current_row), 'Payroll Period', header)
		sheet.write('H'+str(current_row), 'Tax Type', header)
		sheet.write('I'+str(current_row), 'Tax Location', header)
		#sheet.write('J'+str(current_row), 'GAJI POKOK', cell_format)

		# create dictionary array with key
		column_dict 		= {}
		used_component 		= {}
		setup_component		= {}
		
		# kasusnya jika nggak ada opsi payroll periode
		currect_col 	= 0
		current_col2 	= 9

		

		if category_info.code =='DEFAULT':
			sheet.write(column_list[currect_col]+str(current_row), 'Gaji Pokok', header)
			sheet.set_column(current_col2, current_col2, 20)

			column_dict['MTHBASIC'] 			= currect_col
			column_dict['MTH_MITRA_1_BASIC'] 	= currect_col
			column_dict['MTH_OPERATOR_BASIC'] 	= currect_col
			column_dict['MTH_MITRA_2_BASIC'] 	= currect_col
			column_dict['MTH_STAFF_BASIC'] 		= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Acting Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALACT'] 				= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Adjutment', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALADJ'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'All Backup', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALBKUP'] = currect_col
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'BPJS Kes 4% Perusahaan', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_STAFF_ALBPJSKESCOMP'] = currect_col
			column_dict['MTH_OPERATOR_ALBPJSKESCOMP'] = currect_col
			column_dict['MTH_MITRA_1_ALBPJSKESCOMP'] = currect_col
			column_dict['MTH_MITRA_2_ALBPJSKESCOMP'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'BPJS Pensiun 2% Perusahaan', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_STAFF_ALBPJSPENSION'] = currect_col
			column_dict['MTH_MITRA_1_ALBPJSPENSION'] = currect_col
			column_dict['MTH_MITRA_2_ALBPJSPENSION'] = currect_col
			column_dict['MTH_OPERATOR_ALBPJSPENSION'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Competency Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALCOMP'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Guaranteed Incentive', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALGUARAN'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Housing Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALHOME'] = currect_col
			column_dict['ALHOUSEOPS'] = currect_col
			column_dict['ALHOMEOPR'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'ALNETRAL', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALNETRAL'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Handphone Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALHPOPS'] = currect_col
			column_dict['ALHP'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Ins. Libur Nasional', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['Alinslin'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Insentif', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALINS'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'JHT 3.7%', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_OPERATOR_ALJHT'] = currect_col
			column_dict['MTH_MITRA_1_ALJHT'] = currect_col
			column_dict['MTH_MITRA_2_ALJHT'] = currect_col
			column_dict['MTH_STAFF_ALJHT'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'JKK 0.24%', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_MITRA_2_ALJKK'] = currect_col
			column_dict['MTH_MITRA_1_ALJKK'] = currect_col
			column_dict['MTH_MITRA_2_ALJKK'] = currect_col
			column_dict['MTH_STAFF_ALJKK'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'JKM 0.3%', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_MITRA_2_ALJKM'] = currect_col
			column_dict['MTH_MITRA_1_ALJKM'] = currect_col
			column_dict['MTH_MITRA_2_ALJKM'] = currect_col
			column_dict['MTH_STAFF_ALJKM'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Location Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALLOC'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Long Shift', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALLONG'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Location Allowance (TSM)', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALLOCTSM'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Position Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALPOSOPR'] = currect_col
			column_dict['ALPOS'] = currect_col
			setup_component[currect_col] = 1
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Rapel', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALRAPEL'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Meal Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALMEAL'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			

			sheet.write(column_list[currect_col]+str(current_row), 'Travel Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALTRAV'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Transport Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALTRANS'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'TOR Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALTOR'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Performance Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALPERFORM'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan Advance/Pot.Koperasi', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DEADV'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Bonus', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['BONUS'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Car Allowance', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['CARALL'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Pot. Lebih Bayar All', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DELBA'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan Uang Makan/Pot Medical', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DEMEAL'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Uang Jaminan Seragam', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DESRGM'] = currect_col
			column_dict['DESERAGAM'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Pot BPJS Kes 1% Karyawan', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_MITRA_2_DEBPJSKS'] 	= currect_col
			column_dict['MTH_OPERATOR_DEBPJSKS'] 	= currect_col
			column_dict['MTH_STAFF_DEBPJSKS'] 		= currect_col
			column_dict['MTH_MITRA_1_DEBPJSKS'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'BPJS Iuran Pensiun 1%', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MTH_OPERATOR_DEBPJSPENSIONEMP'] = currect_col
			column_dict['MTH_MITRA_2_DEBPJSPENSIONEMP'] = currect_col
			column_dict['MTH_MITRA_1_DEBPJSPENSIONEMP'] = currect_col
			column_dict['MTH_STAFF_DEBPJSPENSIONEMP'] = currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Uang Jaminan Training', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DETRAIN'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Uang Jaminan Training Sec. Guard', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DETRAIN2'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan CAR', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DECAR'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan Asrama', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DEDORMIT'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan Kehadiran', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DEHDR'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan Parkir Kendaraan', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DEPOTMOB'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Potongan Rapel', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['DERAPEL'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Incentive Return', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['INSRETURN'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Medical Reimburshment', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['MEDREIM'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Overtime (Jam)', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['OVTINDEX'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Overtime', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['ALOVT'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1
			

			sheet.write(column_list[currect_col]+str(current_row), 'Subtotal Earning', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['BRUTOINCOME'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Subtotal Deduction', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['BRUTODEDUCTION'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'PPH21', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['PPH21'] 	= currect_col
			setup_component[currect_col] = 1
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Take Home Pay', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['NETPAY'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1
		else:
			sheet.write(column_list[currect_col]+str(current_row), 'THR', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['THR'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'PPH21', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['PPH21'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1

			sheet.write(column_list[currect_col]+str(current_row), 'Take Home Pay', header)
			sheet.set_column(current_col2, current_col2, 20)
			column_dict['NETPAY'] 	= currect_col
			setup_component[currect_col] = 1
			currect_col 	= currect_col + 1
			current_col2 	= current_col2 + 1



		current_row 	= current_row + 1

		# using option
		domain = []

		if year != False:
			domain.append(('year','=', year))

		if month != False:
			domain.append(('month','=', month))

		if category != False:
			domain.append(('struct_id.periode_id.category_id.id','=', category))

		if payroll_periode != False:
			domain.append(('payroll_periode','=', payroll_periode))

		if cost_center != False:
			domain.append(('area_id','=', cost_center))

		if work_location != False:
			domain.append(('employee_id.work_location_id.id','=', work_location))

		if pay_freq != False:
			domain.append(('pay_freq','=', pay_freq))

		if tax_location != False:
			domain.append(('tax_location_id','=', tax_location))

		if tax_type != False:
			domain.append(('tax_type','=', tax_type))
		

		payslip_info = self.env['hr.payslip'].sudo().search(domain)

		counter = 1

		for payslip in payslip_info:
			sheet.write('A'+str(current_row), counter, cell_center)
			sheet.write('B'+str(current_row), payslip.employee_id.nip, cell_left)
			sheet.write('C'+str(current_row), payslip.employee_id.name, cell_left)
			sheet.write('D'+str(current_row), payslip.position_id.name, cell_left)
			sheet.write('E'+str(current_row), payslip.area_id.name, cell_left)
			sheet.write('F'+str(current_row), payslip.payfreq, cell_center)
			sheet.write('G'+str(current_row), payslip.payroll_periode.name, cell_left)
			sheet.write('H'+str(current_row), payslip.tax_type, cell_center)
			sheet.write('I'+str(current_row), payslip.tax_location_id.name, cell_left)
			#sheet.write('J'+str(current_row), 'GAJI POKOK', cell_format)

			# kasusnya jika nggak ada opsi payroll periode
			currect_col 	= 0
			current_col2 	= 9

			if category_info.code =='DEFAULT':
				# based on line ID
				for line in payslip.line_ids:
					if line.code in column_dict:
						sheet.write(column_list[column_dict[line.code]]+str(current_row), line.amount, cell_right_number)

						if  column_dict[line.code] not in used_component:
							used_component[column_dict[line.code]] = 1 
			else:
				for line in payslip.line_ids:
					if line.code in column_dict:
						sheet.write(column_list[column_dict[line.code]]+str(current_row), line.amount, cell_right_number)

						if  column_dict[line.code] not in used_component:
							used_component[column_dict[line.code]] = 1 

			current_row 	= current_row + 1


		# hide unnecessary column
		for setu in setup_component:
			if setu not in used_component:
				sheet.set_column(column_list[setu]+':'+column_list[setu], None, None, {'hidden': True})
		
		workbook.close()
		output.seek(0)
		response.stream.write(output.read())
		output.close()



class FilterEmployee(models.TransientModel):
	_name = "filter.employee.wizard"
	

	name				= fields.Char(string='Nama Pegawai')
	nik					= fields.Char(string='NIK')
	payroll_periode		= fields.Many2one('hr.employee.status',string="Payroll Periode")
	cost_center			= fields.Many2one('area',string="Cost Center")
	work_location		= fields.Many2one('hr.work.location',string="Work Location")
	pay_freq			= fields.Selection([('DAILY','DAILY'),('MONTHLY','MONTHLY')], string='Pay Freq')

	divisi				= fields.Many2one('divisi',string="Divisi")
	cabang				= fields.Many2one('cabang',string="Branch")
	department			= fields.Many2one('hr.department',string="Organization Unit")


	def filter_employee_save(self):
		current_uid = self.env.uid
		#search_info = self.env['ir.ui.view'].sudo().search([('name','=','hr.payslip.custom2.model.view.search')])
		
		# updating filter first to match filtering
		#raise UserError(str(current_uid))
		wizard_info = self.env['filter.employee.wizard'].sudo().search([('create_uid','=',current_uid)], order="create_uid desc", limit=1)

		if wizard_info.id != False:
			wizard_info.write(
				{
					'name'					: self.name,
					'payroll_periode'		: self.payroll_periode.id,
					'cost_center'			: self.cost_center.id,
					'work_location'			: self.work_location.id,
					'pay_freq'				: self.pay_freq,
					'nik'					: self.nik,
					'divisi'				: self.divisi.id,
					'cabang'				: self.cabang.id,
					'department'			: self.department.id
				}
			)
		else:
			self.env['filter.employee.wizard'].create({
				'name'					: self.name,
				'payroll_periode'		: self.payroll_periode.id,
				'cost_center'			: self.cost_center.id,
				'work_location'			: self.work_location.id,
				'pay_freq'				: self.pay_freq,
				'nik'					: self.nik,
				'divisi'				: self.divisi.id,
				'cabang'				: self.cabang.id,
				'department'			: self.department.id
			})

		# name based on filter
		string_filter = ''

		domain = []

		if self.name != False:
			string_filter = string_filter + ' '+self.name
			domain.append(('name','LIKE',self.name))

		if self.payroll_periode.id != False:
			string_filter = string_filter + ' '+ self.payroll_periode.name
			domain.append(('employee_status_id','=',self.payroll_periode.id))

		if self.cost_center.id != False:
			string_filter = string_filter + ' '+ self.cost_center.name
			domain.append(('area','=',self.cost_center.id))

		if self.work_location.id != False:
			string_filter = string_filter + ' '+ self.work_location.name
			domain.append(('work_location_id','=',self.work_location.id))

		if self.pay_freq != False:
			string_filter = string_filter + ' '+ self.pay_freq
			domain.append(('custom3','=',self.pay_freq))

		if self.nik != False:
			string_filter = string_filter + ' '+ self.nik
			domain.append(('nip','=',self.nik))

		if self.divisi.id != False:
			string_filter = string_filter + ' '+ self.divisi.name
			domain.append(('divisi','=',self.divisi.id))

		if self.cabang.id != False:
			string_filter = string_filter + ' '+ self.cabang.name
			domain.append(('cabang','=',self.cabang.id))

		if self.department.id != False:
			string_filter = string_filter + ' '+ self.department.name
			domain.append(('department_id','=',self.department.id))

		
		return {
			'type'					: 'ir.actions.act_window',
			'name'					: 'Employee Data '+string_filter,
			'res_model'				: 'hr.employee',
			'view_id'				: False,
			'views'					: [(False, 'tree'),(False, 'kanban'),(False,'form')],
			'view_type' 			: 'tree',
			'view_mode' 			: 'tree,kanban,form',
			'target' 				: 'main',
			'nodestroy' 			: False,
			#'search_view_id'      	: search_info.id,
			'domain'				: domain
			#'context'             	: { 'search_default_payroll_advanced': 1 }
		}

	
class FilterGeneratePayslip(models.TransientModel):
	_name = "filter.hr.payslip.generate.wizard"
	

	name				= fields.Char(string='Nama Pegawai')
	nik					= fields.Char(string='NIK')
	payroll_periode		= fields.Many2one('hr.employee.status',string="Payroll Periode")
	cost_center			= fields.Many2one('area',string="Cost Center")
	work_location		= fields.Many2one('hr.work.location',string="Work Location")
	pay_freq			= fields.Selection([('DAILY','DAILY'),('MONTHLY','MONTHLY')], string='Pay Freq')

	def filter_payslip_generate_save(self):
		current_uid = self.env.uid
		active_id 	= self.env.context.get('active_id')

		#raise UserError(str(active_id))
		#search_info = self.env['ir.ui.view'].sudo().search([('name','=','hr.payslip.custom2.model.view.search')])
		
		# updating filter first to match filtering
		#raise UserError(str(current_uid))
		wizard_info = self.env['filter.hr.payslip.generate.wizard'].sudo().search([('create_uid','=',current_uid)], order="create_uid desc", limit=1)

		if wizard_info.id != False:
			wizard_info.write(
				{
					'name'					: self.name,
					'payroll_periode'		: self.payroll_periode.id,
					'cost_center'			: self.cost_center.id,
					'work_location'			: self.work_location.id,
					'pay_freq'				: self.pay_freq,
					'nik'					: self.nik
				}
			)
		else:
			self.env['filter.hr.payslip.generate.wizard'].create({
				'name'					: self.name,
				'payroll_periode'		: self.payroll_periode.id,
				'cost_center'			: self.cost_center.id,
				'work_location'			: self.work_location.id,
				'pay_freq'				: self.pay_freq,
				'nik'					: self.nik
			})

		# name based on filter
		string_filter = ''

		domain = []

		if self.name != False:
			string_filter = string_filter + ' '+self.name
			domain.append(('name','LIKE',self.name))

		if self.payroll_periode.id != False:
			string_filter = string_filter + ' '+ self.payroll_periode.name
			domain.append(('employee_status_id','=',self.payroll_periode.id))

		if self.cost_center.id != False:
			string_filter = string_filter + ' '+ self.cost_center.name
			domain.append(('area','=',self.cost_center.id))

		if self.work_location.id != False:
			string_filter = string_filter + ' '+ self.work_location.name
			domain.append(('work_location_id','=',self.work_location.id))

		if self.pay_freq != False:
			string_filter = string_filter + ' '+ self.pay_freq
			domain.append(('custom3','=',self.pay_freq))

		if self.nik != False:
			string_filter = string_filter + ' '+ self.nik
			domain.append(('nip','=',self.nik))


		domain.append(('active','=',True))
		employee_selected = self.env['hr.employee'].sudo().search(domain)
		emp_list = []

		for emp in employee_selected:
			emp_list.append(emp.id)

		if len(emp_list) > 0:
			payslips 	= self.env['hr.payslip']
			[data] 		= self.read()
			active_id 	= self.env.context.get('active_id')
			
			if active_id:
				[run_data] = self.env['hr.payslip.run'].browse(active_id).read(['date_start', 'date_end', 'credit_note','year','month','periode_id'])
				from_date 	= run_data.get('date_start')
				to_date 	= run_data.get('date_end')

			# should validate to protect only one month process
			not_processed 	= self.env['hr.payslip'].sudo().search_read([('employee_id','in',emp_list),('state','=','draft'),('periode_search','!=', run_data.get('month')+'-'+str(run_data.get('year')))],['id','employee_id','position_id','location_id','area_id'])
		
			run_data2 		= self.env['hr.payslip.run'].sudo().search([('id','=',active_id)])
			arr_invalid = [(5, 0, 0)]

			run_data2.write({
				'invalid_ids' : arr_invalid
			})

			if len(not_processed) > 0:
				# nggak bisa
				

				arr_invalid = []

				for notpro in not_processed:
					arr_invalid.append((0,0, {
						'name'          : notpro['employee_id'][0],
						'batch'         : active_id,
						'payslip'       : notpro['id'],
						'position_id'   : notpro['position_id'][0],
						'location_id'   : notpro['location_id'][0],
						'area_id'       : notpro['area_id'][0]
					}))

				run_data2.write({
					'invalid_ids' : arr_invalid
				})


				return {'type': 'ir.actions.act_window_close'}
			else:
				for employee in self.env['hr.employee'].browse(emp_list):
					# from date and to date will be based on hr periode
					slip_data2  = self.env['hr.payslip'].onchange_periode_id(run_data.get('month'), run_data.get('year'), employee.id, run_data.get('periode_id')[0])
					#slip_data   = self.env['hr.payslip'].onchange_employee_id(from_date, to_date, employee.id, contract_id=False)
					
					# adding data based on
					slip_data   = self.env['hr.payslip'].onchange_employee_id2(slip_data2['value'].get('date_from'), slip_data2['value'].get('date_to'),slip_data2['value'].get('attend_start'), slip_data2['value'].get('attend_end'), employee.id, contract_id=False)
					

					res = {
						'employee_id'           : employee.id,
						'name'                  : slip_data['value'].get('name'),
						'struct_id'             : slip_data2['value'].get('struct_id'),
						'contract_id'           : slip_data2['value'].get('contract_id'),
						'payslip_run_id'        : active_id,
						'input_line_ids'        : [(0, 0, x) for x in slip_data['value'].get('input_line_ids')],
						'worked_days_line_ids'  : [(0, 0, x) for x in slip_data['value'].get('worked_days_line_ids')],
						'date_from'             : slip_data2['value'].get('date_from'),
						'date_to'               : slip_data2['value'].get('date_to'),
						'credit_note'           : run_data.get('credit_note'),
						'company_id'            : employee.company_id.id,
						'year'                  : slip_data2['value'].get('year'),
						'month'                 : slip_data2['value'].get('month'),
						'attend_start'          : slip_data2['value'].get('attend_start'),
						'attend_end'            : slip_data2['value'].get('attend_end'),
						'tax'                   : slip_data2['value'].get('tax'),
						'paydate'               : slip_data2['value'].get('paydate'),
						'parent_id'             : False

					}
					payslips += self.env['hr.payslip'].create(res)
				
				payslips.compute_sheet()
				
			
		
		return {'type': 'ir.actions.act_window_close'}
