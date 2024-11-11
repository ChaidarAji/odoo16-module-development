import logging
import pytz
import math

from collections import namedtuple, defaultdict

from datetime import datetime, timedelta, time
from pytz import timezone, UTC
from odoo import models, fields, api, _
from odoo.exceptions import UserError

import openpyxl
import base64
from io import BytesIO

from dateutil.relativedelta import relativedelta
from odoo.tools import float_is_zero
from odoo.tools import date_utils
import io
import json

import requests

try:
   from odoo.tools.misc import xlsxwriter
except ImportError:
   import xlsxwriter


_logger = logging.getLogger(__name__)


class PayslipMonthlyReportWizard(models.TransientModel):
	_name 			= "hr.payslip.monthly.report.wizard"
	_description    = "Payroll Monthly Report"

	name 		= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	year 		= fields.Integer(string='Tahun', Default  = int(datetime.now().strftime("%Y")))
	month    	= fields.Selection(
		[ 
			('01', 'January'),
			('02', 'February'),
			('03', 'March'),
			('04', 'April'),
			('05', 'May'),
			('06', 'June'),
			('07', 'July'),
			('08', 'August'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'December')
		], 
		string   = 'Month',    
		required =True, 
		Default  = datetime.now().strftime("%m"))
	
	cost_center	= fields.Many2one('area',string="Cost Center")
	divisi		= fields.Many2one('divisi',string="Divisi")


	def button_monthly_report(self):
		if self.year == False:
			year 	= int(datetime.now().strftime("%Y"))
		else:
			year 	= self.year

		if self.year == False:
			month 	= datetime.now().strftime("%m")
		else:
			month 	= self.month

		if self.name.id == False:
			name 	= False
		else:
			name 	= self.name.id

		if self.cost_center.id == False:
			cost_center 	= False
		else:
			cost_center 	= self.cost_center.id

		if self.divisi.id == False:
			divisi 	= False
		else:
			divisi 	= self.divisi.id

		data = {
		   'year'					: year,
		   'month'					: month,
		   'periode_id'				: name,
		   'cost_center'			: cost_center,
		   'divisi'					: divisi
	   	}

		return {
		   'type'	: 'ir.actions.report',
		   'data'	: 
		   			{
						'model'				: 'hr.payslip.monthly.report.wizard',
						'options'			: json.dumps(data,default=date_utils.json_default),
						'output_format'		: 'xlsx',
						'report_name'		: 'Monthly Report',
					},
		   'report_type': 'xlsx',
	   }
	
	def get_excel_payroll_monthly_reports(self, data, response):
		year 					= data['year']
		month 					= data['month']
		periode_id 				= data['periode_id']
		cost_center 			= data['cost_center']
		divisi 					= data['divisi']

		output 					= io.BytesIO()
		workbook 				= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 					= workbook.add_worksheet('Rekap Bulanan')

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center', 'valign' : 'vcenter'})
		cell_format.set_text_wrap()
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'valign' : 'vcenter'})
		head.set_text_wrap()
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 12, 'border' : 1, 'valign' : 'vcenter'})
		header.set_text_wrap()
		cell_left 				= workbook.add_format({'font_size': 12, 'align': 'left', 'border' : 1, 'valign' : 'vcenter'})
		cell_left.set_text_wrap()
		cell_center 			= workbook.add_format({'font_size': 12, 'align': 'center', 'border' : 1, 'valign' : 'vcenter'})
		cell_center.set_text_wrap()
		cell_right 				= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'valign' : 'vcenter'})
		cell_right.set_text_wrap()
		cell_right_number 		= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'num_format': '#,##0', 'valign' : 'vcenter'})
		cell_right_number.set_text_wrap()
		cell_left_bold 					= workbook.add_format({'font_size': 12, 'align': 'left', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_left_bold.set_text_wrap()
		cell_center_bold 				= workbook.add_format({'font_size': 12, 'align': 'center', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_center_bold.set_text_wrap()
		cell_right_bold 				= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_right_bold.set_text_wrap()
		cell_right_number_bold 			= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'num_format': '#,##0', 'bold' : True, 'valign' : 'vcenter'})
		cell_right_number_bold.set_text_wrap()

		cell_left_no_border				= workbook.add_format({'font_size': 12, 'align': 'left', 'valign' : 'vcenter'})
		cell_left_no_border.set_text_wrap()
		cell_left_no_border_bold		= workbook.add_format({'font_size': 12, 'align': 'left', 'bold' : True, 'valign' : 'vcenter'})
		cell_left_no_border_bold.set_text_wrap()

		if month =='01':
			nama_bulan = 'JANUARI'
		elif month =='02':
			nama_bulan = 'FEBRUARI'
		elif month =='03':
			nama_bulan = 'MARET'
		elif month =='04':
			nama_bulan = 'APRIL'
		elif month =='05':
			nama_bulan = 'MEI'
		elif month =='06':
			nama_bulan = 'JUNI'
		elif month =='07':
			nama_bulan = 'JULI'
		elif month =='08':
			nama_bulan = 'AGUSTUS'
		elif month =='09':
			nama_bulan = 'SEPTEMBER'
		elif month =='10':
			nama_bulan = 'OKTOBER'
		elif month =='11':
			nama_bulan = 'NOVEMBER'
		elif month =='12':
			nama_bulan = 'DESEMBER'

		sheet.merge_range('A1:U1','PT. Virtus Facility Services',head) 
		sheet.merge_range('A2:U2','Monthly Report',head) 
		
		# divisi tahun bulan
		current_row = 3

		if divisi != False:
			divisi_info = self.env['divisi'].sudo().search([('id','=', divisi)])

			if divisi_info.id != False:
				sheet.merge_range('A'+str(current_row)+':U'+str(current_row),divisi_info.name,head) 
				current_row = current_row + 1

		# tahun bulan
		sheet.merge_range('A'+str(current_row)+':U'+str(current_row),nama_bulan +' '+str(year),head) 
		current_row = current_row + 1
		

		#+str(year)+' BULAN '+nama_bulan, head)
		column_list = [
						'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
					 ]
		
		sheet.set_column(0, 0, 50)
		sheet.set_column(1, 1, 20)
		sheet.set_column(2, 2, 25)
		sheet.set_column(3, 3, 20)
		sheet.set_column(4, 4, 20)
		sheet.set_column(5, 5, 20)
		sheet.set_column(6, 6, 20)
		sheet.set_column(7, 7, 20)
		sheet.set_column(8, 8, 20)
		sheet.set_column(9, 9, 20)
		sheet.set_column(10, 10, 20)
		sheet.set_column(11, 11, 20)
		sheet.set_column(12, 12, 20)
		sheet.set_column(13, 13, 20)
		sheet.set_column(14, 14, 20)
		sheet.set_column(15, 15, 20)
		sheet.set_column(16, 16, 20)
		sheet.set_column(17, 17, 20)
		sheet.set_column(18, 18, 20)
		sheet.set_column(19, 19, 20)
		sheet.set_column(20, 20, 20)


		current_row = current_row + 1
		index = 0
		sheet.write(column_list[index]+str(current_row), 'Area Name', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Cost Center', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Grading', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Man Power', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Salary', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Fixed Allowance', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'variable Allowance', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Comp BPJS TK', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'BPJS Pensiun 2% Perush', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'BPJS Kes 4% Perush', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Overtime', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Total Income', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Comp Iuran BPJS TK', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'BPJS Pensiun 2% Perush2', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'BPJS Kes 4% Perush2', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'JHT Emp 2%, Pot BPJS Kes 1% Kary. BPJS Iuran Pensiun 1%', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'UJS dan Pot Sepatu', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Pot Advance/Koperasi, Pot Kehadiran, Pot Rapel, Pot Lebih Bayar, Uang Jaminan Training, Pot Uang Makan/Medical', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Total Deduction', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Tax', header)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'Net Salary', header)

		current_row = current_row + 1

		# custom query
		query = "SELECT b.id AS area_id, b.name AS area_name, c.code AS code, a.grading, SUM(c.amount) AS jumlah, COUNT(DISTINCT a.employee_id) AS jumlah_pegawai,b.code_site_card AS area_code,d.code AS category FROM hr_payslip a LEFT JOIN area b ON b.id =a.area_id LEFT JOIN hr_payslip_line c ON c.slip_id = a.id LEFT JOIN hr_salary_rule e ON e.id = c.salary_rule_id LEFT JOIN hr_salary_rule_category d ON d.id = e.category_id "
		
		# based on filter
		query = query + "WHERE (a.state = 'done' OR a.state = 'draft') "
		query = query + " AND a.year =" + str(year)
		query = query + " AND month = '"+str(month)+"'"

		if divisi != False:
			query = query + " AND a.divisi_id = "+str(divisi)+" "

		if cost_center != False:
			query = query + " AND a.area_id = "+str(cost_center)+" "

		if periode_id != False:
			query = query + " AND a.periode_id = "+str(periode_id)+" "


		query = query + " GROUP BY b.id, a.grading ,c.code, d.code ORDER BY b.id, a.grading, c.code"
		self.env.cr.execute(query)
		data_result = self.env.cr.fetchall()

		# creating object
		data_dict 		= {}
		data_dict_total = {}

		for pay_line in data_result:
			if str(pay_line[0]) in data_dict_total.keys():
				halo = 1
				#total_sebelum = data_dict_total[str(pay_line[0])]['jumlah_pegawai']
				#data_dict_total[str(pay_line[0])].update({'jumlah_pegawai' : total_sebelum + pay_line[5]})
			else:
				data_dict_total[str(pay_line[0])] = {'jumlah_pegawai' : 0, 'basic' : 0, 'fixed' : 0, 'variable' : 0, 'bpjs_tk' : 0 , 'pensiun_2' : 0, 'bpjs_kes4' : 0, 'overtime' : 0, 'potongan_asuransi' : 0, 'sepatu' : 0, 'potongan_lain' : 0, 'tax' : 0, 'net' : 0}

			if str(pay_line[0])+'-'+str(pay_line[3]) in data_dict.keys():
				halo = 1
			else:
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])] = {'area_id' : pay_line[0], 'area_name' : pay_line[1], 'nominal' : 0 , 'grading' : pay_line[3], 'jumlah_pegawai' : pay_line[5], 'area_code' : pay_line[6], 'category' : pay_line[7], 'basic' : 0, 'fixed' : 0, 'variable' : 0, 'bpjs_tk' : 0 , 'pensiun_2' : 0, 'bpjs_kes4' : 0, 'overtime' : 0, 'potongan_asuransi' : 0, 'sepatu' : 0, 'potongan_lain' : 0, 'tax' : 0, 'net' : 0}
			
			if pay_line[2] =='MTHBASIC':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'basic' : pay_line[4]})
				total_sebelum = data_dict_total[str(pay_line[0])]['basic']
				data_dict_total[str(pay_line[0])].update({'basic' : total_sebelum + pay_line[4] })
			
			if pay_line[7] =='ALW' or pay_line[7] =='ALW_NON_TAX':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['fixed'] = data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['fixed'] +  pay_line[4]
				
				total_sebelum = data_dict_total[str(pay_line[0])]['fixed']
				data_dict_total[str(pay_line[0])].update({'fixed' : total_sebelum + pay_line[4] })

			
			
			if pay_line[7] =='IRREG' or pay_line[7] =='IRREG_NON_TAX':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['variable'] = data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['variable'] +  pay_line[4]

				total_sebelum = data_dict_total[str(pay_line[0])]['variable']
				data_dict_total[str(pay_line[0])].update({'variable' : total_sebelum + pay_line[4] })

			
			if pay_line[2] =='ALBPJSKESCOMP' or pay_line[2] =='ALBPJSKESCOMP_M' or pay_line[2] =='MTH_MITRA_1_ALBPJSKESCOMP' or pay_line[2] =='MTH_MITRA_2_ALBPJSKESCOMP' or pay_line[2] =='MTH_STAFF_ALBPJSKESCOMP' or pay_line[2] =='MTH_OPERATOR_ALBPJSKESCOMP':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'bpjs_kes4' : pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['bpjs_kes4']
				data_dict_total[str(pay_line[0])].update({'bpjs_kes4' : total_sebelum + pay_line[4] })

			
			if pay_line[2] =='ALBPJSPENSION' or pay_line[2] =='ALBPJSPENSION_M' or pay_line[2] =='MTH_MITRA_1_ALBPJSPENSION' or pay_line[2] =='MTH_OPERATOR_ALBPJSPENSION' or pay_line[2] =='MTH_STAFF_ALBPJSPENSION' or pay_line[2] =='MTH_MITRA_2_ALBPJSPENSION':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'pensiun_2' : pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['pensiun_2']
				data_dict_total[str(pay_line[0])].update({'pensiun_2' : total_sebelum + pay_line[4] })

			
			if pay_line[2] =='ALJHT' or pay_line[2] =='ALJHT_M' or pay_line[2] =='ALJKK' or pay_line[2] =='ALJKK_M' or pay_line[2] =='MTH_MITRA_2_ALJHT' or pay_line[2] =='MTH_OPERATOR_ALJHT' or pay_line[2] =='MTH_MITRA_1_ALJHT' or pay_line[2] =='MTH_STAFF_ALJHT' or pay_line[2] =='MTH_STAFF_ALJKK' or pay_line[2] =='MTH_MITRA_1_ALJKK' or pay_line[2] =='MTH_OPERATOR_ALJKK' or pay_line[2] =='MTH_MITRA_2_ALJKK' or pay_line[2] =='MTH_STAFF_ALJKM' or pay_line[2] =='MTH_MITRA_2_ALJKM' or pay_line[2] =='MTH_OPERATOR_ALJKM' or pay_line[2] =='MTH_MITRA_1_ALJKM':
				nilai_sebelum = data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['bpjs_tk']
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'bpjs_tk' : nilai_sebelum + pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['bpjs_tk']
				data_dict_total[str(pay_line[0])].update({'bpjs_tk' : total_sebelum + pay_line[4] })

			
			if pay_line[2] =='ALOVT':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'overtime' : pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['overtime']
				data_dict_total[str(pay_line[0])].update({'overtime' : total_sebelum + pay_line[4] })

			
			#JHT Emp 2%, Pot BPJS Kes 1% Kary. BPJS Iuran Pensiun 1%
			if pay_line[2] =='DEBPJSKS' or pay_line[2] =='DEBPJSKS_M' or pay_line[2] =='DEBPJSKSCOMP' or pay_line[2] =='DEBPJSKSCOMP_M' or pay_line[2] =='DEBPJSPENSION' or pay_line[2] =='DEBPJSPENSION_M' or pay_line[2] =='DEBPJSPENSIONEMP' or pay_line[2] =='DEBPJSPENSIONEMP_M' or pay_line[2] =='MTH_OPERATOR_DEBPJSKS' or pay_line[2] =='MTH_MITRA_2_DEBPJSKS' or pay_line[2] =='MTH_MITRA_1_DEBPJSKS' or pay_line[2] =='MTH_STAFF_DEBPJSKS' or pay_line[2] =='MTH_MITRA_2_PALBPJSKESCOMP' or pay_line[2] =='MTH_MITRA_1_PALBPJSKESCOMP' or pay_line[2] =='MTH_OPERATOR_PALBPJSKESCOMP' or pay_line[2] =='MTH_STAFF_PALBPJSKESCOMP' or pay_line[2] =='MTH_MITRA_2_PALBPJSPENSION' or pay_line[2] =='MTH_OPERATOR_PALBPJSPENSION' or pay_line[2] =='MTH_MITRA_1_PALBPJSPENSION' or pay_line[2] =='MTH_STAFF_PALBPJSPENSION' or pay_line[2] =='MTH_OPERATOR_DEBPJSPENSIONEMP' or pay_line[2] =='MTH_STAFF_DEBPJSPENSIONEMP' or pay_line[2] =='MTH_MITRA_2_DEBPJSPENSIONEMP' or pay_line[2] =='MTH_MITRA_1_DEBPJSPENSIONEMP':
				nilai_sebelum = data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['potongan_asuransi']
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'potongan_asuransi' : nilai_sebelum + pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['potongan_asuransi']
				data_dict_total[str(pay_line[0])].update({'potongan_asuransi' : total_sebelum + pay_line[4] })
			
			if pay_line[2] =='DESPT':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'sepatu' : pay_line[4]})

				total_sebelum = data_dict_total[str(pay_line[0])]['sepatu']
				data_dict_total[str(pay_line[0])].update({'sepatu' : total_sebelum + pay_line[4] })
			
			
			if pay_line[2] =='DECAR' or pay_line[2] =='DEADV' or pay_line[2] =='DEMEAL' or pay_line[2] =='DEDORMIT' or pay_line[2] =='DEHDR' or pay_line[2] =='DEPOTMOB' or pay_line[2] =='DERAPEL' or pay_line[2] =='DESERAGAM' or pay_line[2] =='DESRGM' or pay_line[2] =='DETRAIN' or pay_line[2] =='DETRAIN2' or pay_line[2] =='DELBA':
				nilai_sebelum = data_dict[str(pay_line[0])+'-'+str(pay_line[3])]['potongan_lain']
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'potongan_lain' : nilai_sebelum + pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['potongan_lain']
				data_dict_total[str(pay_line[0])].update({'potongan_lain' : total_sebelum + pay_line[4] })
			
			
			if pay_line[2] =='PPH21':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'tax' : pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['tax']
				data_dict_total[str(pay_line[0])].update({'tax' : total_sebelum + pay_line[4] })
			
			if pay_line[2] =='NETPAY':
				data_dict[str(pay_line[0])+'-'+str(pay_line[3])].update({'net' : pay_line[4]})
		
				total_sebelum = data_dict_total[str(pay_line[0])]['net']
				data_dict_total[str(pay_line[0])].update({'net' : total_sebelum + pay_line[4] })
			

		# writing
		current_area_id = 0
		is_area_change  = False
		for res in self.web_progress_iter(data_dict,"Generate Report"):
			# create payslip

			pegawai_sebelum  = data_dict_total[str(data_dict[res]['area_id'])]['jumlah_pegawai']
			data_dict_total[str(data_dict[res]['area_id'])].update({'jumlah_pegawai' : pegawai_sebelum + data_dict[res]['jumlah_pegawai']})

			if current_area_id == 0:
				current_area_id = data_dict[res]['area_id']
				is_area_change = True
			elif current_area_id != data_dict[res]['area_id']:
				is_area_change = True
				# calculate
				#current_row = current_row
				index = 0
				sheet.write(column_list[index]+str(current_row), '', cell_left)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), '', cell_left)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), 'TOTAL PEGAWAI', cell_left)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['jumlah_pegawai'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['basic'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['fixed'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['variable'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_tk'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['pensiun_2'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['overtime'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['basic'] + data_dict_total[str(current_area_id)]['fixed'] + data_dict_total[str(current_area_id)]['variable'] + data_dict_total[str(current_area_id)]['bpjs_tk'] + data_dict_total[str(current_area_id)]['pensiun_2'] + data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_tk'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['pensiun_2'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['potongan_asuransi'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['sepatu'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['potongan_lain'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row),  data_dict_total[str(current_area_id)]['bpjs_tk'] +data_dict_total[str(current_area_id)]['pensiun_2'] + data_dict_total[str(current_area_id)]['bpjs_kes4'] + data_dict_total[str(current_area_id)]['potongan_asuransi'] + data_dict_total[str(current_area_id)]['sepatu'] + data_dict_total[str(current_area_id)]['potongan_lain'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['tax'], cell_right_number)
				index = index + 1
				sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['net'], cell_right_number)
				index = index + 1

				current_row = current_row + 1
				current_area_id = data_dict[res]['area_id']
			else:
				is_area_change = False

			

			index = 0

			if is_area_change == True:
				sheet.write(column_list[index]+str(current_row), data_dict[res]['area_name'], cell_left)
			else:
				sheet.write(column_list[index]+str(current_row), '', cell_left)
			
			
			index = index + 1
			if is_area_change == True:
				sheet.write(column_list[index]+str(current_row), data_dict[res]['area_code'], cell_left)
			else:
				sheet.write(column_list[index]+str(current_row), '', cell_left)

			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['grading'], cell_left)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['jumlah_pegawai'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['basic'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['fixed'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['variable'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['bpjs_tk'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['pensiun_2'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['bpjs_kes4'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['overtime'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['basic'] + data_dict[res]['fixed'] + data_dict[res]['variable'] + data_dict[res]['bpjs_tk'] +data_dict[res]['pensiun_2'] + data_dict[res]['bpjs_kes4'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['bpjs_tk'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['pensiun_2'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['bpjs_kes4'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['potongan_asuransi'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['sepatu'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['potongan_lain'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row),  data_dict[res]['bpjs_tk'] + data_dict[res]['pensiun_2'] + data_dict[res]['bpjs_kes4'] + data_dict[res]['potongan_asuransi'] + data_dict[res]['sepatu'] + data_dict[res]['potongan_lain'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['tax'], cell_right_number)
			index = index + 1
			sheet.write(column_list[index]+str(current_row), data_dict[res]['net'], cell_right_number)
			index = index + 1
			
			current_row = current_row + 1

		# last
		index = 0
		sheet.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), 'TOTAL PEGAWAI', cell_left)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['jumlah_pegawai'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['basic'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['fixed'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['variable'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_tk'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['pensiun_2'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['overtime'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['basic'] + data_dict_total[str(current_area_id)]['fixed'] + data_dict_total[str(current_area_id)]['variable'] + data_dict_total[str(current_area_id)]['bpjs_tk'] + data_dict_total[str(current_area_id)]['pensiun_2'] + data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_tk'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['pensiun_2'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['potongan_asuransi'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['sepatu'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['potongan_lain'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row),  data_dict_total[str(current_area_id)]['bpjs_tk'] +data_dict_total[str(current_area_id)]['pensiun_2'] + data_dict_total[str(current_area_id)]['bpjs_kes4'] + data_dict_total[str(current_area_id)]['potongan_asuransi'] + data_dict_total[str(current_area_id)]['sepatu'] + data_dict_total[str(current_area_id)]['potongan_lain'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['tax'], cell_right_number)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_total[str(current_area_id)]['net'], cell_right_number)
		index = index + 1

		current_row = current_row + 1		
			
		data_dict_grand = {'jumlah_pegawai' : 0, 'basic' : 0, 'fixed' : 0, 'variable' : 0, 'bpjs_tk' : 0 , 'pensiun_2' : 0, 'bpjs_kes4' : 0, 'overtime' : 0, 'potongan_asuransi' : 0, 'sepatu' : 0, 'potongan_lain' : 0, 'tax' : 0, 'net' : 0}
		for res in self.web_progress_iter(data_dict_total,"Generate Report (CALCULATE TOTAL)"):
			total_sebelum = data_dict_grand['jumlah_pegawai']
			data_dict_grand['jumlah_pegawai'] = total_sebelum + data_dict_total[res]['jumlah_pegawai']

			total_sebelum = data_dict_grand['basic']
			data_dict_grand['basic'] = total_sebelum + data_dict_total[res]['basic']

			total_sebelum = data_dict_grand['fixed']
			data_dict_grand['fixed'] = total_sebelum + data_dict_total[res]['fixed']

			total_sebelum = data_dict_grand['variable']
			data_dict_grand['variable'] = total_sebelum + data_dict_total[res]['variable']

			total_sebelum = data_dict_grand['bpjs_tk']
			data_dict_grand['bpjs_tk'] = total_sebelum + data_dict_total[res]['bpjs_tk']

			total_sebelum = data_dict_grand['pensiun_2']
			data_dict_grand['pensiun_2'] = total_sebelum + data_dict_total[res]['pensiun_2']

			total_sebelum = data_dict_grand['bpjs_kes4']
			data_dict_grand['bpjs_kes4'] = total_sebelum + data_dict_total[res]['bpjs_kes4']

			total_sebelum = data_dict_grand['overtime']
			data_dict_grand['overtime'] = total_sebelum + data_dict_total[res]['overtime']

			total_sebelum = data_dict_grand['potongan_asuransi']
			data_dict_grand['potongan_asuransi'] = total_sebelum + data_dict_total[res]['potongan_asuransi']

			total_sebelum = data_dict_grand['sepatu']
			data_dict_grand['sepatu'] = total_sebelum + data_dict_total[res]['sepatu']

			total_sebelum = data_dict_grand['potongan_lain']
			data_dict_grand['potongan_lain'] = total_sebelum + data_dict_total[res]['potongan_lain']

			total_sebelum = data_dict_grand['tax']
			data_dict_grand['tax'] = total_sebelum + data_dict_total[res]['tax']

			total_sebelum = data_dict_grand['net']
			data_dict_grand['net'] = total_sebelum + data_dict_total[res]['net']

		index = 0
		sheet.write(column_list[index]+str(current_row), 'GRAND TOTAL', cell_left_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), '', cell_left_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), '', cell_left_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['jumlah_pegawai'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['basic'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['fixed'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['variable'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['bpjs_tk'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['pensiun_2'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['bpjs_kes4'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['overtime'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['basic'] + data_dict_total[str(current_area_id)]['fixed'] + data_dict_total[str(current_area_id)]['variable'] + data_dict_total[str(current_area_id)]['bpjs_tk'] + data_dict_total[str(current_area_id)]['pensiun_2'] + data_dict_total[str(current_area_id)]['bpjs_kes4'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['bpjs_tk'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['pensiun_2'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['bpjs_kes4'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['potongan_asuransi'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['sepatu'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['potongan_lain'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row),  data_dict_grand['bpjs_tk'] +data_dict_total[str(current_area_id)]['pensiun_2'] + data_dict_total[str(current_area_id)]['bpjs_kes4'] + data_dict_total[str(current_area_id)]['potongan_asuransi'] + data_dict_total[str(current_area_id)]['sepatu'] + data_dict_total[str(current_area_id)]['potongan_lain'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['tax'], cell_right_number_bold)
		index = index + 1
		sheet.write(column_list[index]+str(current_row), data_dict_grand['net'], cell_right_number_bold)
		index = index + 1

		current_row = current_row + 1	

		# prepared by
		context = self._context
		current_uid = context.get('uid')

		employee_data = self.env['hr.employee'].sudo().search([('user_id','=', current_uid)])
		
		current_row = current_row + 1	
		current_row = current_row + 1	
		sheet.write('F'+str(current_row), 'Prepared By', cell_left_no_border_bold)
		current_row = current_row + 1	
		sheet.write('F'+str(current_row), employee_data.name, cell_left_no_border_bold)
		current_row = current_row + 1	
		sheet.write('F'+str(current_row), employee_data.master_id.name, cell_left_no_border_bold)
		index = index + 1


		sheet2 	= workbook.add_worksheet('Rincian Bulanan')
		sheet2.set_column(0, 0, 10)
		sheet2.set_column(1, 1, 30)
		sheet2.set_column(2, 2, 20)
		sheet2.set_column(3, 3, 20)
		sheet2.set_column(4, 4, 40)
		sheet2.set_column(5, 5, 20)
		sheet2.set_column(6, 6, 40)
		sheet2.set_column(7, 7, 50)
		sheet2.set_column(8, 8, 30)
		sheet2.set_column(9, 9, 20)
		sheet2.set_column(10, 10, 20)
		sheet2.set_column(11, 11, 20)

		sheet2.set_column(12, 12, 30)
		sheet2.set_column(13, 13, 15)

		
		sheet2.set_column(14, 14, 20)
		sheet2.set_column(15, 15, 20)
		sheet2.set_column(16, 16, 20)
		sheet2.set_column(17, 17, 20)
		sheet2.set_column(18, 18, 20)
		sheet2.set_column(19, 19, 20)
		sheet2.set_column(20, 20, 20)

		sheet2.set_column(21, 21, 20)
		sheet2.set_column(22, 22, 20)
		sheet2.set_column(23, 23, 20)
		sheet2.set_column(24, 24, 20)
		sheet2.set_column(25, 25, 20)
		sheet2.set_column(26, 26, 20)
		sheet2.set_column(27, 27, 20)
		sheet2.set_column(28, 28, 20)
		sheet2.set_column(29, 29, 20)
		sheet2.set_column(30, 30, 20)

		sheet2.set_column(31, 31, 20)
		sheet2.set_column(32, 32, 20)
		sheet2.set_column(33, 33, 20)
		sheet2.set_column(34, 34, 20)
		sheet2.set_column(35, 35, 20)
		sheet2.set_column(36, 36, 20)
		sheet2.set_column(37, 37, 20)
		sheet2.set_column(38, 38, 20)
		sheet2.set_column(39, 39, 20)
		sheet2.set_column(40, 40, 20)


		current_row = 1
		index 		= 0
		sheet2.write(column_list[index]+str(current_row), 'No', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Name', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Employee No', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Gov Tax File No', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Position', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Grading', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Departement', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Cost Center', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Grade', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Employment Status', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Join Date', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Termination Date', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Length Of Service', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Tax Status', header)

		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Salary', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Fixed Allowance', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'variable Allowance', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Comp BPJS TK', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'BPJS Pensiun 2% Perush', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'BPJS Kes 4% Perush', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Overtime', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Total Income', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Comp Iuran BPJS TK', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'BPJS Pensiun 2% Perush2', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'BPJS Kes 4% Perush2', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'JHT Emp 2%, Pot BPJS Kes 1% Kary. BPJS Iuran Pensiun 1%', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'UJS dan Pot Sepatu', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Pot Advance/Koperasi, Pot Kehadiran, Pot Rapel, Pot Lebih Bayar, Uang Jaminan Training, Pot Uang Makan/Medical', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Total Deduction', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Tax', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Net Salary', header)

		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'NEJHT', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'NE_BPJS_KS', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'NE_BPJS_PENSIONWAGE', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'NE_KLSRAWAT', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Prorate New Join', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Pay Freq', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Hari Kerja', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'Total Overtime', header)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'NBASICSALARY', header)
		
		current_row = current_row + 1

		# custom query
		query = "SELECT a.id, a.employee_id, f.name AS employee_name, f.nip AS nik, f.npwp_nomor AS npwp, g.name AS jabatan, a.grading, h.name AS department, b.id AS area_id, b.name AS area_name, i.name AS grade, j.name AS contract_type, f.join_date, f.end_date, a.service_length, l.name AS ptkp, a.payfreq, c.code AS code, c.amount AS jumlah, b.code_site_card AS area_code, d.code AS category FROM hr_payslip a LEFT JOIN area b ON b.id =a.area_id LEFT JOIN hr_payslip_line c ON c.slip_id = a.id LEFT JOIN hr_salary_rule e ON e.id = c.salary_rule_id LEFT JOIN hr_salary_rule_category d ON d.id = e.category_id LEFT JOIN hr_employee f ON f.id = a.employee_id LEFT JOIN master_job g ON g.id = a.position_id LEFT JOIN hr_department h ON h.id = a.departement_id LEFT JOIN hr_grade i ON i.id = a.grade_id LEFT JOIN hr_contract k ON k.id = a.contract_id LEFT JOIN hr_contract_type j ON j.id = k.contract_type_id LEFT JOIN hr_ptkp l ON l.id = a.ptkp_id "
		
		# based on filter
		query = query + "WHERE (a.state = 'done' OR a.state = 'draft') "
		query = query + " AND a.year =" + str(year)
		query = query + " AND a.month = '"+str(month)+"'"

		if divisi != False:
			query = query + " AND a.divisi_id = "+str(divisi)+" "

		if cost_center != False:
			query = query + " AND a.area_id = "+str(cost_center)+" "

		if periode_id != False:
			query = query + " AND a.periode_id = "+str(periode_id)+" "


		self.env.cr.execute(query)
		data_result = self.env.cr.fetchall()

		no = 1
		data_dict 		= {}
		data_dict_total =  { 'basic' : 0, 'fixed' : 0, 'variable' : 0, 'bpjs_tk' : 0 , 'pensiun_2' : 0, 'bpjs_kes4' : 0, 'overtime' : 0, 'potongan_asuransi' : 0, 'sepatu' : 0, 'potongan_lain' : 0, 'tax' : 0, 'net' : 0, 'jam_overtime' : 0,  'ne_jht' : 0, 'ne_bpjs_ks' : 0, 'ne_bpjs_pensiun' : 0}

		collected_ids = []


		for pay_line in data_result:
			collected_ids.append(pay_line[0])

			if str(pay_line[0]) in data_dict.keys():
				halo = 1
			else:
				data_dict[str(pay_line[0])] = {'employee_id' : pay_line[1], 'employee_name' : pay_line[2], 'employee_nik' : pay_line[3], 'npwp' : pay_line[4], 'jabatan' : pay_line[5], 'grading' : pay_line[6], 'department' : pay_line[7], 'area_id' : pay_line[8], 'area_name' : pay_line[9], 'grade' : pay_line[10], 'contract_type' : pay_line[11], 'join_date' : pay_line[12], 'end_date' : pay_line[13], 'service_length' : pay_line[14], 'ptkp' : pay_line[15], 'payfreq' : pay_line[16], 'area_code' : pay_line[19] , 'basic' : 0, 'fixed' : 0, 'variable' : 0, 'bpjs_tk' : 0 , 'pensiun_2' : 0, 'bpjs_kes4' : 0, 'overtime' : 0, 'potongan_asuransi' : 0, 'sepatu' : 0, 'potongan_lain' : 0, 'tax' : 0, 'net' : 0, 'ne_jht' : 0, 'ne_bpjs_ks' : 0, 'ne_bpjs_pensiun' : 0, 'kls_rawat' : '', 'prorate' : '',  'hari_kerja' : 0, 'jam_overtime' : 0}
		
			
			if pay_line[17] =='BASICBPJS':
				data_dict[str(pay_line[0])].update({'ne_bpjs_ks' : pay_line[18]})

				total_sebelum = data_dict_total['ne_bpjs_ks']
				data_dict_total.update({'ne_bpjs_ks' : total_sebelum + pay_line[18] })


			if pay_line[17] =='NEJHT':
				data_dict[str(pay_line[0])].update({'ne_jht' : pay_line[18]})

				total_sebelum = data_dict_total['ne_jht']
				data_dict_total.update({'ne_jht' : total_sebelum + pay_line[18] })

			if pay_line[17] =='NE_BPJS_PENSIONWAGE':
				data_dict[str(pay_line[0])].update({'ne_bpjs_pensiun' : pay_line[18]})

				total_sebelum = data_dict_total['ne_bpjs_pensiun']
				data_dict_total.update({'ne_bpjs_pensiun' : total_sebelum + pay_line[18] })


			if pay_line[17] =='NOTPRORATE' and pay_line[18] == '1':
				data_dict[str(pay_line[0])].update({'prorate' : '0'})
			elif pay_line[17] =='NOTPRORATE' and pay_line[18] == '0':
				data_dict[str(pay_line[0])].update({'prorate' : '1'})

			if pay_line[17] =='OVTINDEX':
				data_dict[str(pay_line[0])].update({'jam_overtime' : pay_line[18]})
			
			if pay_line[17] =='MTHBASIC':
				data_dict[str(pay_line[0])].update({'basic' : pay_line[18]})
				total_sebelum = data_dict_total['basic']
				data_dict_total.update({'basic' : total_sebelum + pay_line[18] })
			
			if pay_line[20] =='ALW' or pay_line[20] =='ALW_NON_TAX':
				data_dict[str(pay_line[0])]['fixed'] = data_dict[str(pay_line[0])]['fixed'] +  pay_line[18]
				
				total_sebelum = data_dict_total['fixed']
				data_dict_total.update({'fixed' : total_sebelum + pay_line[18] })
			
			if pay_line[20] =='IRREG' or pay_line[20] =='IRREG_NON_TAX':
				data_dict[str(pay_line[0])]['variable'] = data_dict[str(pay_line[0])]['variable'] +  pay_line[18]

				total_sebelum = data_dict_total['variable']
				data_dict_total.update({'variable' : total_sebelum + pay_line[18] })

			
			if pay_line[17] =='ALBPJSKESCOMP' or pay_line[17] =='ALBPJSKESCOMP_M' or pay_line[17] =='MTH_MITRA_1_ALBPJSKESCOMP' or pay_line[17] =='MTH_MITRA_2_ALBPJSKESCOMP' or pay_line[17] =='MTH_STAFF_ALBPJSKESCOMP' or pay_line[17] =='MTH_OPERATOR_ALBPJSKESCOMP':
				data_dict[str(pay_line[0])].update({'bpjs_kes4' : pay_line[18]})
		
				total_sebelum = data_dict_total['bpjs_kes4']
				data_dict_total.update({'bpjs_kes4' : total_sebelum + pay_line[18] })

			
			if pay_line[17] =='ALBPJSPENSION' or pay_line[17] =='ALBPJSPENSION_M' or pay_line[17] =='MTH_MITRA_1_ALBPJSPENSION' or pay_line[17] =='MTH_OPERATOR_ALBPJSPENSION' or pay_line[17] =='MTH_STAFF_ALBPJSPENSION' or pay_line[17] =='MTH_MITRA_2_ALBPJSPENSION':
				data_dict[str(pay_line[0])].update({'pensiun_2' : pay_line[18]})
		
				total_sebelum = data_dict_total['pensiun_2']
				data_dict_total.update({'pensiun_2' : total_sebelum + pay_line[18] })

			
			if pay_line[17] =='ALJHT' or pay_line[17] =='ALJHT_M' or pay_line[17] =='ALJKK' or pay_line[17] =='ALJKK_M' or pay_line[17] =='MTH_MITRA_2_ALJHT' or pay_line[17] =='MTH_OPERATOR_ALJHT' or pay_line[17] =='MTH_MITRA_1_ALJHT' or pay_line[17] =='MTH_STAFF_ALJHT' or pay_line[17] =='MTH_STAFF_ALJKK' or pay_line[17] =='MTH_MITRA_1_ALJKK' or pay_line[17] =='MTH_OPERATOR_ALJKK' or pay_line[17] =='MTH_MITRA_2_ALJKK' or pay_line[17] =='MTH_STAFF_ALJKM' or pay_line[17] =='MTH_MITRA_2_ALJKM' or pay_line[17] =='MTH_OPERATOR_ALJKM' or pay_line[17] =='MTH_MITRA_1_ALJKM':
				nilai_sebelum = data_dict[str(pay_line[0])]['bpjs_tk']
				data_dict[str(pay_line[0])].update({'bpjs_tk' : nilai_sebelum + pay_line[18]})
		
				total_sebelum = data_dict_total['bpjs_tk']
				data_dict_total.update({'bpjs_tk' : total_sebelum + pay_line[18] })

			
			if pay_line[17] =='ALOVT':
				data_dict[str(pay_line[0])].update({'overtime' : pay_line[18]})
		
				total_sebelum = data_dict_total['overtime']
				data_dict_total.update({'overtime' : total_sebelum + pay_line[18] })

			
			#JHT Emp 2%, Pot BPJS Kes 1% Kary. BPJS Iuran Pensiun 1%
			if pay_line[17] =='DEBPJSKS' or pay_line[17] =='DEBPJSKS_M' or pay_line[17] =='DEBPJSKSCOMP' or pay_line[17] =='DEBPJSKSCOMP_M' or pay_line[17] =='DEBPJSPENSION' or pay_line[17] =='DEBPJSPENSION_M' or pay_line[17] =='DEBPJSPENSIONEMP' or pay_line[17] =='DEBPJSPENSIONEMP_M' or pay_line[17] =='MTH_OPERATOR_DEBPJSKS' or pay_line[17] =='MTH_MITRA_2_DEBPJSKS' or pay_line[17] =='MTH_MITRA_1_DEBPJSKS' or pay_line[17] =='MTH_STAFF_DEBPJSKS' or pay_line[17] =='MTH_MITRA_2_PALBPJSKESCOMP' or pay_line[17] =='MTH_MITRA_1_PALBPJSKESCOMP' or pay_line[17] =='MTH_OPERATOR_PALBPJSKESCOMP' or pay_line[17] =='MTH_STAFF_PALBPJSKESCOMP' or pay_line[17] =='MTH_MITRA_2_PALBPJSPENSION' or pay_line[17] =='MTH_OPERATOR_PALBPJSPENSION' or pay_line[17] =='MTH_MITRA_1_PALBPJSPENSION' or pay_line[17] =='MTH_STAFF_PALBPJSPENSION' or pay_line[17] =='MTH_OPERATOR_DEBPJSPENSIONEMP' or pay_line[17] =='MTH_STAFF_DEBPJSPENSIONEMP' or pay_line[17] =='MTH_MITRA_2_DEBPJSPENSIONEMP' or pay_line[17] =='MTH_MITRA_1_DEBPJSPENSIONEMP':
				nilai_sebelum = data_dict[str(pay_line[0])]['potongan_asuransi']
				data_dict[str(pay_line[0])].update({'potongan_asuransi' : nilai_sebelum + pay_line[18]})
		
				total_sebelum = data_dict_total['potongan_asuransi']
				data_dict_total.update({'potongan_asuransi' : total_sebelum + pay_line[18] })
			
			if pay_line[17] =='DESPT':
				data_dict[str(pay_line[0])].update({'sepatu' : pay_line[18]})

				total_sebelum = data_dict_total['sepatu']
				data_dict_total.update({'sepatu' : total_sebelum + pay_line[18] })
			
			
			if pay_line[17] =='DECAR' or pay_line[17] =='DEADV' or pay_line[17] =='DEMEAL' or pay_line[17] =='DEDORMIT' or pay_line[17] =='DEHDR' or pay_line[17] =='DEPOTMOB' or pay_line[17] =='DERAPEL' or pay_line[17] =='DESERAGAM' or pay_line[17] =='DESRGM' or pay_line[17] =='DETRAIN' or pay_line[17] =='DETRAIN2' or pay_line[17] =='DELBA':
				nilai_sebelum = data_dict[str(pay_line[0])]['potongan_lain']
				data_dict[str(pay_line[0])].update({'potongan_lain' : nilai_sebelum + pay_line[18]})
		
				total_sebelum = data_dict_total['potongan_lain']
				data_dict_total.update({'potongan_lain' : total_sebelum + pay_line[18] })
			
			
			if pay_line[17] =='PPH21':
				data_dict[str(pay_line[0])].update({'tax' : pay_line[18]})
		
				total_sebelum = data_dict_total['tax']
				data_dict_total.update({'tax' : total_sebelum + pay_line[18] })
			
			if pay_line[17] =='NETPAY':
				data_dict[str(pay_line[0])].update({'net' : pay_line[18]})
		
				total_sebelum = data_dict_total['net']
				data_dict_total.update({'net' : total_sebelum + pay_line[18] })
			
		
		# hitung hari kerja 
		collected_ids_join = ','.join(str(x) for x in collected_ids)
		query = "SELECT payslip_id, SUM(number_of_days) AS jumlah FROM hr_payslip_worked_days a WHERE a.payslip_id IN ("+collected_ids_join+") GROUP BY a.payslip_id"
		self.env.cr.execute(query)
		hari_kerja_list = self.env.cr.fetchall()

		for pay_line in hari_kerja_list:
			data_dict[str(pay_line[0])].update({'hari_kerja' : pay_line[1]})
		
		no = 1
		for res in self.web_progress_iter(data_dict,"Generate Report Employee"):
			index 		= 0
			sheet2.write(column_list[index]+str(current_row), no, cell_center)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['employee_name'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['employee_nik'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['npwp'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['jabatan'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['grading'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['department'],cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['area_name'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['grade'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['contract_type'], cell_left)
			index = index + 1
			
			# join date and end date problem, convert to string
			year2 		= data_dict[res]['join_date'].year
			month2		= str(data_dict[res]['join_date'].month).zfill(2)
			hari2		= str(data_dict[res]['join_date'].day).zfill(2)

			join_date	= str(year2)+'-'+month2+'-'+hari2

			end_date = ''

			if data_dict[res]['end_date'] != False and data_dict[res]['end_date'] != None:
				year2 		= data_dict[res]['end_date'].year
				month2		= str(data_dict[res]['end_date'].month).zfill(2)
				hari2		= str(data_dict[res]['end_date'].day).zfill(2)

				end_date	= str(year2)+'-'+month2+'-'+hari2


			sheet2.write(column_list[index]+str(current_row), join_date, cell_center)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), end_date, cell_center)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['service_length'],cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['ptkp'], cell_left)

			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['basic'],cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row),data_dict[res]['fixed'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['variable'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['bpjs_tk'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['pensiun_2'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['bpjs_kes4'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['overtime'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['basic'] + data_dict[res]['fixed'] + data_dict[res]['variable'] + data_dict[res]['bpjs_tk'] + data_dict[res]['pensiun_2'] + data_dict[res]['bpjs_kes4'] + data_dict[res]['overtime'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['bpjs_tk'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['pensiun_2'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['bpjs_kes4'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['potongan_asuransi'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row),data_dict[res]['sepatu'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['potongan_lain'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['bpjs_tk'] + data_dict[res]['pensiun_2'] + data_dict[res]['bpjs_kes4'] + data_dict[res]['potongan_asuransi'] + data_dict[res]['sepatu'] +data_dict[res]['potongan_lain'] , cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['tax'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['net'], cell_right_number)

			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['ne_jht'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['ne_bpjs_ks'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['ne_bpjs_pensiun'], cell_right_number)
			index = index + 1

			if data_dict[res]['ne_bpjs_ks'] > 4000000:
				sheet2.write(column_list[index]+str(current_row), '1', header)
			elif data_dict[res]['ne_bpjs_ks'] > 0:
				sheet2.write(column_list[index]+str(current_row), '2', header)
			else:
				sheet2.write(column_list[index]+str(current_row), '', header)

			index = index + 1
			
			sheet2.write(column_list[index]+str(current_row),data_dict[res]['prorate'], cell_center)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['payfreq'], cell_left)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['hari_kerja'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['jam_overtime'], cell_right_number)
			index = index + 1
			sheet2.write(column_list[index]+str(current_row), data_dict[res]['basic'], cell_right_number)
			
			current_row = current_row + 1
			no = no + 1


		# grand total
		index 		= 0
		sheet2.write(column_list[index]+str(current_row), '', cell_center)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), 'TOTAL', cell_left_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '',cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_center)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_center)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '',cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)

		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['basic'],cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['fixed'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['variable'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['bpjs_tk'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['pensiun_2'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['bpjs_kes4'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['overtime'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['basic'] + data_dict_total['fixed'] + data_dict_total['variable'] + data_dict_total['bpjs_tk'] + data_dict_total['pensiun_2'] + data_dict_total['bpjs_kes4'] + data_dict_total['overtime'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['bpjs_tk'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['pensiun_2'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['bpjs_kes4'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['potongan_asuransi'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['sepatu'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['potongan_lain'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['bpjs_tk'] + data_dict_total['pensiun_2'] + data_dict_total['bpjs_kes4'] + data_dict_total['potongan_asuransi'] + data_dict_total['sepatu'] +data_dict_total['potongan_lain'] , cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['tax'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['net'], cell_right_number_bold)

		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['ne_jht'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['ne_bpjs_ks'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['ne_bpjs_pensiun'], cell_right_number_bold)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', header)
		index = index + 1
			
		sheet2.write(column_list[index]+str(current_row), '', cell_center)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_left)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_right_number)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), '', cell_right_number)
		index = index + 1
		sheet2.write(column_list[index]+str(current_row), data_dict_total['basic'], cell_right_number_bold)
			
		current_row = current_row + 1
		
		workbook.close()
		output.seek(0)
		response.stream.write(output.read())
		output.close()



	

class PayslipCashReportWizard(models.TransientModel):
	_name 			= "hr.payslip.cash.report.wizard"
	_description    = "Payroll Gaji Cash"

	name 		= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	year 		= fields.Integer(string='Tahun', Default  = int(datetime.now().strftime("%Y")))
	month    	= fields.Selection(
		[ 
			('01', 'January'),
			('02', 'February'),
			('03', 'March'),
			('04', 'April'),
			('05', 'May'),
			('06', 'June'),
			('07', 'July'),
			('08', 'August'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'December')
		], 
		string   = 'Month',    
		required =True, 
		Default  = datetime.now().strftime("%m"))
	
	cost_center	= fields.Many2one('area',string="Cost Center")

	divisi		= fields.Many2one('divisi',string="Divisi")

	def button_cash_report(self):
		if self.year == False:
			year 	= int(datetime.now().strftime("%Y"))
		else:
			year 	= self.year

		if self.year == False:
			month 	= datetime.now().strftime("%m")
		else:
			month 	= self.month

		if self.name.id == False:
			name 	= False
		else:
			name 	= self.name.id

		if self.cost_center.id == False:
			cost_center 	= False
		else:
			cost_center 	= self.cost_center.id

		if self.divisi.id == False:
			divisi 	= False
		else:
			divisi 	= self.divisi.id

		data = {
		   'year'					: year,
		   'month'					: month,
		   'periode_id'				: name,
		   'cost_center'			: cost_center,
		   'divisi'					: divisi
	   	}

		return {
		   'type'	: 'ir.actions.report',
		   'data'	: 
		   			{
						'model'				: 'hr.payslip.cash.report.wizard',
						'options'			: json.dumps(data,default=date_utils.json_default),
						'output_format'		: 'xlsx',
						'report_name'		: 'Gaji Cash',
					},
		   'report_type': 'xlsx',
	   }
	
	def get_excel_payroll_cash_reports(self, data, response):
		year 					= data['year']
		month 					= data['month']
		periode_id 				= data['periode_id']
		cost_center 			= data['cost_center']
		divisi 					= data['divisi']

		output 					= io.BytesIO()
		workbook 				= xlsxwriter.Workbook(output, {'in_memory': True})

		cell_format 					= workbook.add_format({'font_size': 12, 'align': 'center', 'valign' : 'vcenter'})
		head 							= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'valign' : 'vcenter'})
		header 							= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 12, 'border' : 1, 'valign' : 'vcenter'})
		cell_left 						= workbook.add_format({'font_size': 12, 'align': 'left', 'border' : 1, 'valign' : 'vcenter'})
		cell_center 					= workbook.add_format({'font_size': 12, 'align': 'center', 'border' : 1, 'valign' : 'vcenter'})
		cell_right 						= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'valign' : 'vcenter'})
		cell_right_number 				= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'num_format': '#,##0', 'valign' : 'vcenter'})
		cell_left_bold 					= workbook.add_format({'font_size': 12, 'align': 'left', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_center_bold 				= workbook.add_format({'font_size': 12, 'align': 'center', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_right_bold 				= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_right_number_bold 			= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'num_format': '#,##0', 'bold' : True, 'valign' : 'vcenter'})
		cell_left_no_border				= workbook.add_format({'font_size': 12, 'align': 'left', 'valign' : 'vcenter'})
		cell_left_no_border_bold		= workbook.add_format({'font_size': 12, 'align': 'left', 'bold' : True, 'valign' : 'vcenter'})
		cell_center_no_border_bold		= workbook.add_format({'font_size': 12, 'align': 'center', 'bold' : True, 'valign' : 'vcenter'})
		cell_left_no_border_bold.set_text_wrap()
		cell_format.set_text_wrap()
		head.set_text_wrap()
		header.set_text_wrap()
		cell_right_number_bold.set_text_wrap()
		cell_left_no_border.set_text_wrap()
		cell_right_bold.set_text_wrap()
		cell_center_bold.set_text_wrap()
		cell_left_bold.set_text_wrap()
		cell_right_number.set_text_wrap()
		cell_right.set_text_wrap()
		cell_center.set_text_wrap()
		cell_left.set_text_wrap()
		cell_center_no_border_bold.set_text_wrap()

		if month =='01':
			nama_bulan = 'JANUARI'
		elif month =='02':
			nama_bulan = 'FEBRUARI'
		elif month =='03':
			nama_bulan = 'MARET'
		elif month =='04':
			nama_bulan = 'APRIL'
		elif month =='05':
			nama_bulan = 'MEI'
		elif month =='06':
			nama_bulan = 'JUNI'
		elif month =='07':
			nama_bulan = 'JULI'
		elif month =='08':
			nama_bulan = 'AGUSTUS'
		elif month =='09':
			nama_bulan = 'SEPTEMBER'
		elif month =='10':
			nama_bulan = 'OKTOBER'
		elif month =='11':
			nama_bulan = 'NOVEMBER'
		elif month =='12':
			nama_bulan = 'DESEMBER'



		# based on divisi
		if divisi != False:
			divisi_info = self.env['divisi'].sudo().search([('id','=', divisi)])
		else:
			divisi_info = self.env['divisi'].sudo().search([])

		
		
		sheet_counter 	= 0
		sheet_name		= ['a','b','c','d','e','f','g','h','i','j','k','l']
		sheet_list		= {}

		column_list = [
			'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
			'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
			'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
			'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
			'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
		]


		
		for divi in divisi_info:


			# how to set dynamic variable name
			sheet_list[sheet_name[sheet_counter]] 	= workbook.add_worksheet(divi.name)

			index = 0
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 10)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 25)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 50)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 50)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 50)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 20)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].set_column(index, index, 30)


			sheet_list[sheet_name[sheet_counter]].merge_range('A1:U1','PT. Virtus Facility Services',head) 
			sheet_list[sheet_name[sheet_counter]].merge_range('A2:U2','DATA CASH',head) 
			sheet_list[sheet_name[sheet_counter]].merge_range('A3:U3',divi.name ,head) 
			sheet_list[sheet_name[sheet_counter]].merge_range('A4:U4',nama_bulan +' '+str(year),head) 

			current_row = 5
			

			index = 0
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'No', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'NIK', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Nama', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Join Date', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Emp End Date', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Jabatan', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Area', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Keterangan', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'PIC Hold', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Total', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Tgl Proses', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Potongan Tambahan', header)
			index = index + 1
			sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), 'Nilai Ditransfer', header)

			current_row = current_row + 1

			#select by divisi
			query = "SELECT f.id, f.nip AS nik, f.name AS nama, f.join_date, f.end_date, b.name AS area_name, f.description AS keterangan, g.name AS pic, c.amount, h.name AS jabatan FROM hr_payslip a  JOIN area b ON b.id =a.area_id LEFT JOIN hr_payslip_line c ON c.slip_id = a.id  LEFT JOIN hr_salary_rule e ON e.id = c.salary_rule_id LEFT JOIN hr_employee f ON f.id = a.employee_id LEFT JOIN hr_employee g ON g.id = f.parent_id LEFT JOIN master_job h ON h.id = a.position_id WHERE e.code = 'NETPAY' "
			query = query + "AND (a.state = 'done' OR a.state = 'draft') "
			query = query + " AND a.year =" + str(year)
			query = query + " AND month = '"+str(month)+"'"
			query = query + " AND a.divisi_id = "+str(divi.id)+" "
			
			if cost_center != False:
				query = query + " AND a.area_id = "+str(cost_center)+" "

			if periode_id != False:
				query = query + " AND a.periode_id = "+str(periode_id)+" "

			self.env.cr.execute(query)
			data_result = self.env.cr.fetchall()

			no = 1
			for pay_line in data_result:
				index 		= 0
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), no, cell_center)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), str(pay_line[1]), cell_center)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), str(pay_line[2]), cell_left)
				
				# format tgl
				year2 		= pay_line[3].year
				month2		= str(pay_line[3].month).zfill(2)
				hari2		= str(pay_line[3].day).zfill(2)

				join_date	= str(year2)+'-'+month2+'-'+hari2

				end_date = ''

				if pay_line[4] != False and pay_line[4] != None:
					year2 		= pay_line[4].year
					month2		= str(pay_line[4].month).zfill(2)
					hari2		= str(pay_line[4].day).zfill(2)

					end_date	= str(year2)+'-'+month2+'-'+hari2

				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), join_date, cell_center)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), end_date, cell_center)
				
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), str(pay_line[9]), cell_left)
				
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), str(pay_line[5]), cell_left)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), str(pay_line[6]), cell_left)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), str(pay_line[7]), cell_left)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), pay_line[8], cell_right_number)
			
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row),'', cell_left)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), '', cell_left)
				index = index + 1
				sheet_list[sheet_name[sheet_counter]].write(column_list[index]+str(current_row), '', cell_left)
				
				current_row = current_row + 1
				no = no + 1

				
				
				

			# prepared by
			context = self._context
			current_uid = context.get('uid')

			employee_data = self.env['hr.employee'].sudo().search([('user_id','=', current_uid)])
				
			current_row = current_row + 1	
			current_row = current_row + 1	
			sheet_list[sheet_name[sheet_counter]].write('B'+str(current_row), 'Prepared By', cell_left_no_border_bold)
			sheet_list[sheet_name[sheet_counter]].write('F'+str(current_row), 'Acknowledge By', cell_center_no_border_bold)
			sheet_list[sheet_name[sheet_counter]].write('F'+str(current_row), 'Approved By', cell_center_no_border_bold)
			current_row = current_row + 1	
			sheet_list[sheet_name[sheet_counter]].write('B'+str(current_row), employee_data.name, cell_left_no_border_bold)
			current_row = current_row + 1	
			sheet_list[sheet_name[sheet_counter]].write('B'+str(current_row), employee_data.master_id.name, cell_left_no_border_bold)
			index = index + 1
			
			sheet_counter = sheet_counter + 1

	
		# only can be prepared by
		

		workbook.close()
		output.seek(0)
		response.stream.write(output.read())
		output.close()

class PayslipEmployeeDataReportWizard(models.TransientModel):
	_name 			= "hr.payslip.employee.data.report.wizard"
	_description    = "Payroll Detail Gaji Pegawai"

	name 		= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode", required = True) 
	cost_center	= fields.Many2one('area',string="Cost Center")

	divisi		= fields.Many2one('divisi',string="Divisi")

	def button_employee_data_report(self):
	
		if self.name.id == False:
			name 	= False
		else:
			name 	= self.name.id

		if self.cost_center.id == False:
			cost_center 	= False
		else:
			cost_center 	= self.cost_center.id

		if self.divisi.id == False:
			divisi 	= False
		else:
			divisi 	= self.divisi.id

		data = {
		   'periode_id'				: name,
		   'cost_center'			: cost_center,
		   'divisi'					: divisi
	   	}

		return {
		   'type'	: 'ir.actions.report',
		   'data'	: 
		   			{
						'model'				: 'hr.payslip.employee.data.report.wizard',
						'options'			: json.dumps(data,default=date_utils.json_default),
						'output_format'		: 'xlsx',
						'report_name'		: 'Payslip By Employee Data Report',
					},
		   'report_type': 'xlsx',
	   }
	
	def get_excel_payroll_employee_data_reports(self, data, response):
		periode_id 				= data['periode_id']
		cost_center 			= data['cost_center']
		divisi 					= data['divisi']

		periode_info			= self.env['hr.periode'].sudo().search([('id','=', periode_id)])

		output 					= io.BytesIO()
		workbook 				= xlsxwriter.Workbook(output, {'in_memory': True})

		cell_format 					= workbook.add_format({'font_size': 12, 'align': 'center', 'valign' : 'vcenter'})
		head 							= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16, 'valign' : 'vcenter'})
		header 							= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 12, 'border' : 1, 'valign' : 'vcenter'})
		cell_left 						= workbook.add_format({'font_size': 12, 'align': 'left', 'border' : 1, 'valign' : 'vcenter'})
		cell_center 					= workbook.add_format({'font_size': 12, 'align': 'center', 'border' : 1, 'valign' : 'vcenter'})
		cell_right 						= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'valign' : 'vcenter'})
		cell_right_number 				= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'num_format': '#,##0', 'valign' : 'vcenter'})
		cell_left_bold 					= workbook.add_format({'font_size': 12, 'align': 'left', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_center_bold 				= workbook.add_format({'font_size': 12, 'align': 'center', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_right_bold 				= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'bold' : True, 'valign' : 'vcenter'})
		cell_right_number_bold 			= workbook.add_format({'font_size': 12, 'align': 'right', 'border' : 1, 'num_format': '#,##0', 'bold' : True, 'valign' : 'vcenter'})
		cell_left_no_border				= workbook.add_format({'font_size': 12, 'align': 'left', 'valign' : 'vcenter'})
		cell_left_no_border_bold		= workbook.add_format({'font_size': 12, 'align': 'left', 'bold' : True, 'valign' : 'vcenter'})
		cell_center_no_border_bold		= workbook.add_format({'font_size': 12, 'align': 'center', 'bold' : True, 'valign' : 'vcenter'})
		cell_left_no_border_bold.set_text_wrap()
		cell_format.set_text_wrap()
		head.set_text_wrap()
		header.set_text_wrap()
		cell_right_number_bold.set_text_wrap()
		cell_left_no_border.set_text_wrap()
		cell_right_bold.set_text_wrap()
		cell_center_bold.set_text_wrap()
		cell_left_bold.set_text_wrap()
		cell_right_number.set_text_wrap()
		cell_right.set_text_wrap()
		cell_center.set_text_wrap()
		cell_left.set_text_wrap()
		cell_center_no_border_bold.set_text_wrap()

		column_list = [
			'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
			'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
			'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
			'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
			'BS','BT','BU', 'BV','BW','BX', 'BY','BZ'
		]

		sheet 	= workbook.add_worksheet('Employee Payroll Data Report')
		index = 0
		sheet.set_column(index, index, 10)
		index = index + 1
		sheet.set_column(index, index, 25)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 15)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 25)
		index = index + 1
		sheet.set_column(index, index, 25)
		index = index + 1
		sheet.set_column(index, index, 25)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 25)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)
		index = index + 1
		sheet.set_column(index, index, 20)
		index = index + 1
		sheet.set_column(index, index, 40)

		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 30)
		index = index + 1
		sheet.set_column(index, index, 30)

		# header
		sheet.merge_range('A1:BG1','PT. Virtus Facility Services',cell_left_no_border_bold) 
		sheet.merge_range('A2:BG2','Employee Payroll Data Report',cell_left_no_border_bold) 

		current_row = 4


		index = 0
		current_row_plus = current_row + 1
		current_row_plus2 = current_row + 2
		#sheet.write(column_list[index]+str(current_row), 'No', header)
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "No", header)
		index = index + 1
		#sheet.write(column_list[index]+str(current_row), 'Employee No', header)
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Employee No", header)
		index = index + 1
		#sheet.write(column_list[index]+str(current_row), 'Employee Name', header)
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Employee Name", header)


		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Gender", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Position Name", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Organization Unit", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Join Date", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Termination Date", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Basic Salary", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Efektif Salary Date", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Salary Formula", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Organization Unit", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Salary Received", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Pay Frequency", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Gov. Tax File No.", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Gov. Tax File Date", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Tax Location", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Tax Type", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Tax Status", header)
		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index]+str(current_row_plus2), "Total Dependent", header)

		index = index + 1
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field JKK 0.89 =1 or 1.74 = 2 or 0.54 = 3 or 1.27 = 4", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Housing Allowance (payfield 10)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Daily Allowance (payfield 11)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Potongan Asrama (payfield 12)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Handphone Allowance (payfield 13)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field JKK/JKM PP49 100% [1=yes]", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field BPJS KS Mitra", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field BPJS TK Mitra", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Meal Allowance (payfield 2)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Acting Allowance (payfield 3)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field TOR Allowance (payfield 4)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Long Shift (payfield 5)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Guaranteed Incentive (payfield 6)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2
		
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Location Allowance (payfield 7)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Position Allowance (payfield 8)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Payroll Field Transport Allowance (payfield 9)", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Value", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Effective date", header)
		index = index + 2

		
		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 1]+str(current_row), "Insurance Branch", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Insurance No", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index]+str(current_row_plus2), "Insurance Name", header)
		index = index + 2

		sheet.merge_range(column_list[index]+str(current_row)+':'+column_list[index + 4]+str(current_row), "Payroll Period", header)
		sheet.merge_range(column_list[index]+str(current_row_plus)+':'+column_list[index + 4]+str(current_row_plus), periode_info.name, header)
		sheet.merge_range(column_list[index]+str(current_row_plus2)+':'+column_list[index]+str(current_row_plus2), "Included", header)
		sheet.merge_range(column_list[index + 1]+str(current_row_plus2)+':'+column_list[index + 1]+str(current_row_plus2), "Comp Bank", header)
		sheet.merge_range(column_list[index + 2]+str(current_row_plus2)+':'+column_list[index + 2]+str(current_row_plus2), "Comp Acc", header)
		sheet.merge_range(column_list[index + 3]+str(current_row_plus2)+':'+column_list[index + 3]+str(current_row_plus2), "Emp Bank", header)
		sheet.merge_range(column_list[index + 4]+str(current_row_plus2)+':'+column_list[index + 4]+str(current_row_plus2), "Emp Acc Period Code", header)
		index = index + 3

		# query langsung tanpa split
		query = "SELECT a.nip, a.name, a.gender, b.name AS jabatan, c.name AS department, a.join_date, a.end_date, d.wage, d.date_start, a.custom3, a.npwp_nomor, a.bpjs_kesehatan, e.name AS tax_location, f.name AS tax_type, g.name AS ptkp, l.name AS bank, i.acc_number AS rekening, l.id AS group_bank FROM hr_employee a LEFT JOIN master_job b ON b.id = a.master_id LEFT JOIN hr_department c ON c.id = a.department_id LEFT JOIN hr_contract d ON d.id = a.contract_id LEFT JOIN tax_location e ON e.id = d.tax_location_id LEFT JOIN tax_type f ON f.id = d.tax_type_id LEFT JOIN hr_ptkp g ON g.id = a.ptkp_id LEFT JOIN hr_employee_pembayaran h ON h.name = a.id LEFT JOIN res_partner_bank i ON i.id = h.rekening LEFT JOIN hr_periode_category j ON j.id = h.category LEFT JOIN res_bank k ON k.id = i.bank_id LEFT JOIN res_bank_group l ON l.id = k.group_id LEFT JOIN hr_periode m ON m.category_id = j.id WHERE a.active = True  "
		query = query + " AND m.id =  "+str(periode_id)

		if cost_center != False:
			query = query + " AND a.area = "+str(cost_center)+" "

		if divisi != False:
			query = query + " AND a.divisi = "+str(divisi)+" "

		self.env.cr.execute(query)
		data_result = self.env.cr.fetchall()
		no = 1


		workbook.close()
		output.seek(0)
		response.stream.write(output.read())
		output.close()





	



class PayslipGenerateWizard(models.TransientModel):
	_name 			= "hr.payslip.generate.wizard"
	_description    = "Generate Payslip"
	
	name 				= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	periode_label 		= fields.Char(string='Periode Tanggal')
	area_id 			= fields.Many2one(string = "Cost Center" ,comodel_name = "area") 
	employee_number		= fields.Integer(string='Jumlah Pegawai', readonly=True)
	employee_ids 		= fields.One2many('hr.payslip.generate.employee.wizard', 'name', string='Pegawai Terpilih', readonly=False)
	error_ids 			= fields.One2many('hr.payslip.generate.error.wizard', 'name', string='Error', readonly=False)
	is_error_visible	= fields.Boolean(string='Error Visible', default = False)
	choose_employee		= fields.Boolean(string='Pilih Pegawai', default = False)
	show_employee		= fields.Boolean(string='Tampilkan List Pegawai', default = False)
	processed			= fields.Integer(string='Terproses')
	employee_list 		= fields.One2many('hr.payslip.generate.list.wizard', 'name', string='Daftar Pegawai', readonly=False)


	@api.onchange('show_employee')
	def onchange_show_employee(self):
		# find IDS based on selected 
		if self.name.id != False:
			periode_info = self.env['hr.periode'].sudo().search([('id','=', self.name.id)])

			year 		= periode_info.salaryenddate.year
			month		= str(periode_info.salaryenddate.month).zfill(2)

			query 		= "select a.id, a.name, a.nip, a.contract_id, a.company_id from hr_employee AS a left join hr_payslip as b ON b.employee_id = a.id AND b.year = "+str(year)+" AND b.month = '"+str(month)+"' and b.state not IN ('verify','cancel') WHERE a.employee_status_id = "+str(periode_info.status_id.id)+" AND a.active = True AND b.id IS NULL"
					
			if self.area_id.id != False:
				query = query + ' AND a.area = '+str(self.area_id.id)

			self.env.cr.execute(query)
			data_result 			= self.env.cr.fetchall()
			
			
			# populate IDS
			arr_ids = [(5,0,0)]
			counter = 0
			for tup in data_result:
				counter = counter + 1
				arr_ids.append((0,0,{'employee_id' : tup[0], 'nik' : str(tup[2])}))

			self.employee_number 	= counter

			#res = self.write({
			#	'name' 					: self.name.id,
			#	'periode_label'			: self.periode_label,
			#	'area_id'				: self.area_id.id,
			#	'employee_number'		: self.employee_number,
			#	'employee_ids'			: self.employee_ids,
			#	'error_ids'				: self.error_ids,
			#	'is_error_visible'		: self.is_error_visible,
			#	'choose_employee'		: self.choose_employee,
			#	'show_employee'			: self.show_employee,
			#	'processed'				: self.processed,
			#	'employee_list' 		: arr_ids
			#})

			self.write({
				'employee_list' 	: arr_ids,
				'employee_number'	: counter
			})	

			



	@api.onchange('name','area_id')
	def onchange_name(self):
		if self.name.id != False:
			periode_info = self.env['hr.periode'].sudo().search([('id','=', self.name.id)])

			# converting to date label
			bulan_list = {
				'1' : 'Januari',
				'2' : 'Februari',
				'3' : 'Maret',
				'4' : 'April',
				'5' : 'Mei',
				'6' : 'Juni',
				'7' : 'Juli',
				'8' : 'Agustus',
				'9' : 'September',
				'10' : 'Oktober',
				'11' : 'November',
				'12' : 'Desember'
			}

			self.periode_label = str(periode_info.salarystartdate.day) +' '+ bulan_list[str(periode_info.salarystartdate.month)] + ' '+ str(periode_info.salarystartdate.year) +' s.d. '+str(periode_info.salaryenddate.day) +' '+ bulan_list[str(periode_info.salaryenddate.month)] + ' '+ str(periode_info.salaryenddate.year)

			# ambil tahun
			year 		= periode_info.salaryenddate.year
			month		= str(periode_info.salaryenddate.month).zfill(2)

			#query 		= "select a.id, a.name, a.nip from hr_employee AS a left join hr_payslip as b ON b.employee_id = a.id AND b.year = "+str(year)+" AND b.month = '"+str(month)+"' and b.state not IN ('verify','cancel') WHERE a.employee_status_id = "+str(periode_info.status_id.id)+" AND a.active = True AND b.id IS NULL"
			query 		= "select COUNT(a.id) as jumlah from hr_employee AS a left join hr_payslip as b ON b.employee_id = a.id AND b.year = "+str(year)+" AND b.month = '"+str(month)+"' and b.state not IN ('verify','cancel') WHERE a.employee_status_id = "+str(periode_info.status_id.id)+" AND a.active = True AND b.id IS NULL"
			
			if self.area_id.id != False:
				query = query + ' AND a.area = '+str(self.area_id.id)

			self.env.cr.execute(query)
			data_result = self.env.cr.fetchall()


			self.employee_number = data_result[0][0]
			
			#data_dict	= {}

			#for tup in data_result:
			#	data_dict[tup[0]] = {'id' : tup[0], 'name' : tup[1], 'nip' : tup[2]}
			
			#data_result = dict(data_result)


			#_logger.error('DATA RESULT')
			#_logger.error(data_dict)

	def button_generate_payslip(self):
		# find role
		periode_info = self.env['hr.periode'].sudo().search([('id','=', self.name.id)])

		year 		= periode_info.salaryenddate.year
		month		= str(periode_info.salaryenddate.month).zfill(2)

		self.processed = 0


		emp_ids = []
		if len(self.employee_ids) > 0:
			for emp in self.employee_ids:
				emp_ids.append(emp.employee_id.id)
			
			data_result = self.env['hr.employee'].sudo().search([('id','in', emp_ids)])  #"select a.id, a.name, a.nip, a.contract_id, a.company_id from hr_employee AS a WHERE a.id IN("
			
			# change to dict
			data_dict	= {}
			for dat in data_result:
				data_dict[dat.nip] = {'id' : dat.id, 'name' : dat.name, 'nip' : dat.nip,'contract_id' : dat.contract_id.id, 'company_id' : dat.company_id.id}
		else:
			query 		= "select a.id, a.name, a.nip, a.contract_id, a.company_id, a.end_date from hr_employee AS a left join hr_payslip as b ON b.employee_id = a.id AND b.year = "+str(year)+" AND b.month = '"+str(month)+"' and b.state not IN ('verify','cancel') WHERE a.employee_status_id = "+str(periode_info.status_id.id)+" AND a.active = True AND b.id IS NULL "
				
			if self.area_id.id != False:
				query = query + ' AND a.area = '+str(self.area_id.id)

			self.env.cr.execute(query)
			data_result 			= self.env.cr.fetchall()
			self.employee_number 	= data_result[0][0]

			data_dict	= {}

			for tup in data_result:
				data_dict[str(tup[2])] = {'id' : tup[0], 'name' : tup[1], 'nip' : str(tup[2]),'contract_id' : tup[3], 'company_id' : tup[4], 'end_date' : tup[5]}
				
		#data_result = dict(data_result)
		#if 'DO1913034' in data_dict.keys():
		#	raise UserError('ADA')


		if periode_info.useattend == True:
			# call presence
			headers = {
				"Content-Type"	: "text/plain", 
				"Accept"		: "*/*", 
				"Catch-Control"	: "no-cache",
				"apikey" 		: "ms3Rko81bTrbO85ZCpk691PBaItghIyEbCvw0Ex"
			}

			url 		= "https://api.smartpresence.id/v1/customrequest/virtusabsence?startdate="+datetime.strftime(periode_info.attendstartdate, "%Y-%m-%d")+"&enddate="+datetime.strftime(periode_info.attendenddate, "%Y-%m-%d")
			response 	= requests.get(url, headers=headers)

			json_response = response.json()

			counter_response = 0

			if json_response['status'] == 'OK':
				if len(json_response['data']) > 0 :
					response_dict = {}
					for resp in json_response['data']:
						response_dict[str(resp['employee_number'])] = {'employee_id' : resp['employee_id'], 'employee_number' : str(resp['employee_number']), 'employee_name' : resp['employee_name'], 'workday_count' : resp['workday_count'], 'presence_count' : resp['presence_count']}

					for resp2 in self.web_progress_iter(response_dict, "Pengambilan Presensi"):
						# check if element object exist
						if resp2 in data_dict.keys():
							schedule_info = self.env['hr.payroll.presensi.bulanan'].sudo().search([('name','=', data_dict[resp2]['id']),('year','=', year),('month','=', month)])
							# insert/update to database
							if schedule_info.id == False:
								self.env['hr.payroll.presensi.bulanan'].create({
									'name' 			: data_dict[resp2]['id'],
									'nik'			: response_dict[resp2]['employee_number'],
									'start'			: periode_info.attendstartdate,
									'end'			: periode_info.attendenddate,
									'year'			: year,
									'month'			: month,
									'work_days'		: response_dict[resp2]['workday_count'],
									'work_shift'	: response_dict[resp2]['presence_count']
								})
							else:
								for sche in schedule_info:
									sche.write({
										'name' 			: data_dict[resp2]['id'],
										'nik'			: response_dict[resp2]['employee_number'],
										'start'			: periode_info.attendstartdate,
										'end'			: periode_info.attendenddate,
										'year'			: year,
										'month'			: month,
										'work_days'		: response_dict[resp2]['workday_count'],
										'work_shift'	: response_dict[resp2]['presence_count']
									})

							#total_hari 	= json_response['data'][0]['presence_count']
							#total_shift	= json_response['data'][0]['presence_count']
			else:
				raise UserError('Pengambilan data smart presence gagal')
		

		arr_invalid 	= [(5, 0, 0)]
		error_visible 	= False
		for res in self.web_progress_iter(data_dict,"Generate Payslip"):
			# create payslip
			employee 			= self.env['hr.employee'].browse(data_dict[res]['id'])

			if employee.employment_status.id == False:
				# to error
				error_visible = True
				arr_invalid.append((0,0,{'nik' : employee.nip, 'employee_id' : employee.id, 'description' : 'Employment status belum di set'}))
			elif employee.contract_id.id == False:
				error_visible = True
				arr_invalid.append((0,0,{'nik' : employee.nip, 'employee_id' : employee.id, 'description' : 'Tidak ada kontrak berjalan'}))
			elif employee.end_date != False and employee.end_date < periode_info.salarystartdate:
				error_visible = True
				arr_invalid.append((0,0,{'nik' : employee.nip, 'employee_id' : employee.id, 'description' : 'Pegawai status aktif tapi end date nya kurang dari periode gaji'}))
			else:
				slip_name 			= employee.nip +' '+employee.name+' '+employee.master_id.name + ' '+employee.area.name
				try:
					struktur_info 		= self.env['hr.payroll.structure'].sudo().search([('periode_id','=',periode_info.id)])

					if len(struktur_info) > 0:
						for stru in struktur_info:
							struct_id = stru.id

					# calculate worked days based on presence saved
					if employee.grade_id.category_id.code == 'FRONTLINER':
						grading = 'OPERATOR'
					else:
						grading = 'PENGAWAS'

					payslip = self.env['hr.payslip'].create({
						'employee_id'           : data_dict[res]['id'],
						'name'                  : slip_name,
						'periode_id'			: periode_info.id,
						'struct_id'             : struct_id,
						'contract_id'           : employee.contract_id.id,
						'payslip_run_id'        : False,
						'input_line_ids'        : [],
						'worked_days_line_ids'  : [],
						'date_from'             : periode_info.salarystartdate,
						'date_to'               : periode_info.salaryenddate,
						'credit_note'           : '',
						'company_id'            : employee.company_id.id,
						'year'                  : year,
						'month'                 : month,
						'attend_start'          : periode_info.attendstartdate,
						'attend_end'            : periode_info.attendenddate,
						'tax'                   : periode_info.taxdate,
						'paydate'               : periode_info.paydate,
						'parent_id'             : False,
						'is_active'				: True,
						'grade_id'				: employee.grade_id.id,
						'grading'				: grading
					})

					self.processed = self.processed  + 1

					contracts = self.env['hr.contract'].browse(employee.contract_id.id)
					worked_days_line_ids 	= payslip.get_worked_day_lines(contracts, periode_info.attendstartdate, periode_info.attendenddate)
					input_line_ids 			= payslip.get_inputs(contracts, periode_info.salarystartdate, periode_info.salaryenddate)



					payslip.write({
						'worked_days_line_ids'	: [(0, 0, x) for x in worked_days_line_ids],
						'input_line_ids'		: [(0, 0, x) for x in input_line_ids]
					})

					payslip.compute_sheet()
				except:
					error_visible = True
					arr_invalid.append((0,0,{'nik' : employee.nip, 'employee_id' : employee.id, 'description' : 'Error sistem : Tidak diketahui'}))

					#raise UserError('Ada error perhitungan pada pegawai '+employee.name+' ('+employee.nip+')')

		self.error_ids 			= arr_invalid
		self.is_error_visible 	= error_visible	
		#raise UserError('HALOO')


class PayslipGenerateEmployeeWizard(models.TransientModel):
	_name 			= "hr.payslip.generate.employee.wizard"
	_description    = "Generate Payslip"

	name 			= fields.Many2one(string = "Payslip Wizard" ,comodel_name = "hr.payslip.generate.wizard") 
	employee_id		= fields.Many2one(string = "nama Pegawai" ,comodel_name = "hr.employee")
	nik				= fields.Char(string='NIK', default=lambda self: self.employee_id.nip)

	@api.onchange('employee_id')
	def onchange_employee_id(self):
		if self.employee_id.id != False:
			self.nik	= self.employee_id.nip
	 

class PayslipGenerateErrorWizard(models.TransientModel):
	_name 			= "hr.payslip.generate.error.wizard"
	_description    = "Generate Payslip"

	name 			= fields.Many2one(string = "Payslip Wizard" ,comodel_name = "hr.payslip.generate.wizard") 
	employee_id		= fields.Many2one(string = "nama Pegawai" ,comodel_name = "hr.employee") 
	nik				= fields.Char(string='NIK', default=lambda self: self.employee_id.nip)
	description		= fields.Char(string='Keterangan')

class PayslipGenerateListWizard(models.TransientModel):
	_name 			= "hr.payslip.generate.list.wizard"
	_description    = "Generate Payslip"

	name 			= fields.Many2one(string = "Payslip Wizard" ,comodel_name = "hr.payslip.generate.wizard") 
	employee_id		= fields.Many2one(string = "nama Pegawai" ,comodel_name = "hr.employee")
	nik				= fields.Char(string='NIK', default=lambda self: self.employee_id.nip)
	employee_name	= fields.Char(string='Pegawai', related="employee_id.name")


	def button_selected_employee(self):
		# get wizard parent
		wizard_info = self.env['hr.payslip.generate.wizard'].sudo().search([('id','=', self.name.id)])

		wizard_info.write({
			'employee_ids' : [(0,0,{'employee_id' : self.employee_id.id, 'nik' : self.employee_id.nip})]
		})

class HrJournalEstimasiGajiWizard(models.TransientModel):
	_name 			= "hr.journal.estimasi.gaji.wizard"
	_description    = "Journal Estimasi Gaji"

	def _get_default_bulan(self):
		payslip_id = self.env['hr.payslip'].sudo().search([], order="year desc, month desc", limit=1)
		return payslip_id.month

	def _get_default_tahun(self):
		payslip_id = self.env['hr.payslip'].sudo().search([], order="year desc, month desc", limit=1)
		return payslip_id.year

	name 		= fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	year = fields.Integer(string='Tahun', default=_get_default_tahun)
	month = fields.Selection(
		[ 
			('01', 'Januari'),
			('02', 'Februari'),
			('03', 'Maret'),
			('04', 'April'),
			('05', 'Mei'),
			('06', 'Juni'),
			('07', 'Juli'),
			('08', 'Agustus'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'Desember')
		], 
		string   = 'Bulan',    
		required =True, default=_get_default_bulan)
	cost_center	= fields.Many2one('area',string="Cost Center")
	divisi		= fields.Many2one('divisi',string="Divisi")
	excel_file	= fields.Binary(string="File Excel")

	def button_journal_estimasi_gaji(self):
		output 						= io.BytesIO()
		workbook 					= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 						= workbook.add_worksheet()

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		head_info 				= workbook.add_format({'align': 'left', 'bold': True, 'font_size': 12})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})


		column_list = [
						'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ']
		#data payslip domain
		domain = []
		if self.name:
			domain.append(('periode_id','=', self.name.id))
		if self.cost_center.id != False:
			domain.append(('area_id','=', self.cost_center.id))
		if self.divisi.id != False:
			domain.append(('divisi_id','=', self.divisi.id))
		if self.year:
			domain.append(('year','=', self.year))
		if self.month:
			domain.append(('month','=', self.month))


		# Title
		sheet.merge_range('B1:H1', 'Jurnal Estimasi Gaji', head)
		sheet.write('B2', 'Periode : '+(self.name.name or ''), head_info)
		sheet.write('B3', 'Cost Center : '+(self.cost_center.name or ''), head_info)
		sheet.write('B4', 'Divisi : '+(self.divisi.name or ''), head_info)
		sheet.write('D2', 'Bulan : ' + dict(self._fields['month'].selection).get(self.month, ''), head_info)
		sheet.write('D3', 'Tahun : '+str(self.year), head_info)

		sheet.set_column(0, 0, 5)
		sheet.set_column(1, 20, 20)
		sheet.write(5, 0, 'No.', header)
		sheet.write(5, 1, 'Name', header)
		sheet.write(5, 2, 'Empployee No.', header)
		sheet.write(5, 3, 'Position', header)
		sheet.write(5, 4, 'Cost Center', header)
		sheet.write(5, 5, 'Salary', header)
		sheet.write(5, 6, 'Salary Full', header)
		sheet.write(5, 7, 'Pay Freq.', header)
		sheet.write(5, 8, 'Estimasi Gaji', header)

		# Generate data
		payslip_info = self.env['hr.payslip'].sudo().search(domain)
		print(payslip_info)
		current_row = 6
		no = 0
		total_estimasi = 0
		for slip in payslip_info:
			#compute values
			salary = 0
			salary_full = 0
			salary_id = slip.details_by_salary_rule_category.filtered(lambda x: x.code == 'NETPAY')
			if salary_id:
				salary = salary_id.total
				total_estimasi += salary
			salary_full_id = slip.details_by_salary_rule_category.filtered(lambda x: x.code == 'MTHBASIC')
			if salary_full_id:
				if slip.payfreq == 'MONTHLY':
					salary_full = salary_full_id.total
				else:
					salary_full = salary_full_id.total / 26

			sheet.write(current_row, 0, no+1, cell_center)
			sheet.write(current_row, 1, slip.employee_id.name, cell_left)
			sheet.write(current_row, 2, slip.employee_id.nip, cell_center)
			sheet.write(current_row, 3, slip.position_id.name, cell_left)
			sheet.write(current_row, 4, slip.area_id.name, cell_left)
			sheet.write(current_row, 5, salary, cell_right_number)
			sheet.write(current_row, 6, salary_full, cell_right_number)
			sheet.write(current_row, 7, slip.payfreq, cell_center)
			sheet.write(current_row, 8, salary, cell_right_number)

			current_row += 1
			no += 1	
		sheet.write(current_row, 7, 'Total', cell_center)
		sheet.write(current_row, 8, total_estimasi, cell_right_number)

		workbook.close()
		output.seek(0)
		file_base64 = base64.b64encode(output.read())
		output.close()
		self.excel_file = file_base64
		return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=hr.journal.estimasi.gaji.wizard&field=excel_file&download=true&id=%s&filename=%s' % (self.id, "Journal_Estimasi_Gaji.xlsx"),
            'target': 'self'
        }
	
class HrJournalPengembalianPajakWizard(models.TransientModel):
	_name 			= "hr.journal.pengembalian.pajak.wizard"
	_description    = "Journal Pengembalian Pajak"

	def _get_default_bulan(self):
		payslip_id = self.env['hr.payslip'].sudo().search([], order="year desc, month desc", limit=1)
		return payslip_id.month

	def _get_default_tahun(self):
		payslip_id = self.env['hr.payslip'].sudo().search([], order="year desc, month desc", limit=1)
		return payslip_id.year

	name = fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	year = fields.Integer(string='Tahun', default=_get_default_tahun)
	month = fields.Selection(
		[ 
			('01', 'Januari'),
			('02', 'Februari'),
			('03', 'Maret'),
			('04', 'April'),
			('05', 'Mei'),
			('06', 'Juni'),
			('07', 'Juli'),
			('08', 'Agustus'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'Desember')
		], 
		string   = 'Bulan',    
		required =True, default=_get_default_bulan)
	divisi_ids = fields.Many2many('divisi', string="Divisi")
	excel_file = fields.Binary(string="File Excel")

	def button_journal_pengembalian_pajak(self):
		output 						= io.BytesIO()
		workbook 					= xlsxwriter.Workbook(output, {'in_memory': True})
		sheet 						= workbook.add_worksheet()

		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		head_info 				= workbook.add_format({'align': 'left', 'bold': True, 'font_size': 12})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})


		column_list = [
						'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ']
		#data payslip domain
		domain = []
		if self.name:
			domain.append(('periode_id','=', self.name.id))
		if self.divisi_ids:
			domain.append(('divisi_id','in', self.divisi_ids.ids))
		if self.year:
			domain.append(('year','=', self.year))
		if self.month:
			domain.append(('month','=', self.month))


		# Title
		sheet.merge_range('B1:H1', 'Jurnal Pengembalian Pajak', head)
		sheet.write('B2', 'Periode : '+(self.name.name or ''), head_info)
		sheet.write('B3', 'Divisi : '+(self.divisi_ids.name or ''), head_info)
		sheet.write('D2', 'Bulan : ' + dict(self._fields['month'].selection).get(self.month, ''), head_info)
		sheet.write('D3', 'Tahun : '+str(self.year), head_info)

		sheet.set_column(0, 0, 20)
		sheet.set_column(1, 20, 20)
		sheet.write(5, 0, 'Area Name', header)
		sheet.write(5, 1, 'Cost Center', header)
		sheet.write(5, 2, 'Man Power', header)
		sheet.write(5, 3, 'Tax', header)

		workbook.close()
		output.seek(0)
		file_base64 = base64.b64encode(output.read())
		output.close()
		self.excel_file = file_base64
		return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=hr.journal.pengembalian.pajak.wizard&field=excel_file&download=true&id=%s&filename=%s' % (self.id, "Journal_Pengembalian_Pajak.xlsx"),
            'target': 'self'
        }

class HrJournalDetailGajiWizard(models.TransientModel):
	_name 			= "hr.journal.detail.gaji.wizard"
	_description    = "Journal Detail Gaji"

	def _get_default_bulan(self):
		payslip_id = self.env['hr.payslip'].sudo().search([], order="year desc, month desc", limit=1)
		return payslip_id.month

	def _get_default_tahun(self):
		payslip_id = self.env['hr.payslip'].sudo().search([], order="year desc, month desc", limit=1)
		return payslip_id.year

	name = fields.Many2one(string = "Payroll Periode" ,comodel_name = "hr.periode") 
	year = fields.Integer(string='Tahun', default=_get_default_tahun)
	month = fields.Selection(
		[ 
			('01', 'Januari'),
			('02', 'Februari'),
			('03', 'Maret'),
			('04', 'April'),
			('05', 'Mei'),
			('06', 'Juni'),
			('07', 'Juli'),
			('08', 'Agustus'),
			('09', 'September'),
			('10', 'Oktober'),
			('11', 'November'),
			('12', 'Desember')
		], 
		string   = 'Bulan',    
		required =True, default=_get_default_bulan)
	divisi_ids = fields.Many2many('divisi', string="Divisi")
	excel_file = fields.Binary(string="File Excel")

	def button_journal_detail_gaji(self):
		output 						= io.BytesIO()
		workbook 					= xlsxwriter.Workbook(output, {'in_memory': True})
		#STYLE
		cell_format 			= workbook.add_format({'font_size': 12, 'align': 'center'})
		head 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 16})
		head_info 				= workbook.add_format({'align': 'left', 'bold': True, 'font_size': 12})
		header 					= workbook.add_format({'align': 'center', 'bold': True, 'font_size': 10, 'border' : 1})
		cell_left 				= workbook.add_format({'font_size': 10, 'align': 'left', 'border' : 1})
		cell_center 			= workbook.add_format({'font_size': 10, 'align': 'center', 'border' : 1})
		cell_right 				= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1})
		cell_right_number 		= workbook.add_format({'font_size': 10, 'align': 'right', 'border' : 1, 'num_format': '#,##0'})


		column_list = [
						'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB',
				 		'AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR',
						'AS','AT','AU', 'AV','AW','AX', 'AY','AZ','BA','BB',
				 		'BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR',
						'BS','BT','BU', 'BV','BW','BX', 'BY','BZ']

		#DIVISI
		divisi_ids = self.divisi_ids or self.env['divisi'].search([])
		for divisi in divisi_ids:
			sheet 						= workbook.add_worksheet(name=f"Journal {divisi.name}")

			#data payslip domain
			domain = [('divisi_id','=', divisi.id)]
			if self.name:
				domain.append(('periode_id','=', self.name.id))
			if self.year:
				domain.append(('year','=', self.year))
			if self.month:
				domain.append(('month','=', self.month))

			# Payslip Data
			payslip_ids = self.env['hr.payslip'].sudo().search(domain)

			# Title 
			sheet.merge_range('B1:H1', f'Jurnal {divisi.name}', head)
			sheet.write('B2', 'Periode : '+(self.name.name or ''), head_info)
			sheet.write('D2', 'Bulan : ' + dict(self._fields['month'].selection).get(self.month, ''), head_info)
			sheet.write('D3', 'Tahun : '+str(self.year), head_info)

			# REKAP HEADER
			sheet.set_column(0, 0, 20)
			sheet.set_column(1, 20, 20)
			sheet.write(5, 0, 'Area Name', header)
			sheet.write(5, 1, 'Cost Center', header)
			sheet.write(5, 2, 'Grading', header)
			sheet.write(5, 3, 'Man Power', header)
			sheet.write(5, 4, 'Salary', header)
			sheet.write(5, 5, 'JHT 3.7%', header)
			sheet.write(5, 6, 'JKK 0.25%', header)
			sheet.write(5, 7, 'Meal Allowance', header)
			sheet.write(5, 8, 'Compentency Allowance', header)
			sheet.write(5, 9, 'BPJS Pensiun 2% Perusahaan', header)
			sheet.write(5, 10, 'BPJS Kesehatan 4% Perusahaan', header)
			sheet.write(5, 11, 'Long Shift', header)
			sheet.write(5, 12, 'Adjustment', header)
			sheet.write(5, 13, 'Guaranted Incentive', header)
			sheet.write(5, 14, 'Other Allowance', header)
			sheet.write(5, 15, 'Incentive', header)
			sheet.write(5, 16, 'Location Allowance', header)
			sheet.write(5, 17, 'Position Allowance', header)
			sheet.write(5, 18, 'Transport Allowance', header)
			sheet.write(5, 19, 'Overtime', header)
			sheet.write(5, 20, 'ALL BACKUP', header)
			sheet.write(5, 21, 'Ins. Libur Nasional', header)
			sheet.write(5, 22, 'Daily Allowance', header)
			sheet.write(5, 23, 'Housing Allowance', header)
			sheet.write(5, 24, 'Handphone Allowance', header)
			sheet.write(5, 25, 'Performance Allowance', header)
			sheet.write(5, 26, 'Medical', header)
			sheet.write(5, 27, 'Acting Allowance', header)
			sheet.write(5, 28, 'Total Income', header)
			sheet.write(5, 29, 'Pot JHT 3.7%`', header)
			sheet.write(5, 30, 'JHT employee 2%', header)
			sheet.write(5, 31, 'Pot JKK 0.24%', header)
			sheet.write(5, 32, 'Pot JKM 0.3%', header)
			sheet.write(5, 33, 'BPJS Pensiun 2% Perusahaan2', header)
			sheet.write(5, 34, 'BPJS Kesehatan 4% Perusahaan2', header)
			sheet.write(5, 35, 'Pot BPJS Kes 1% Employee', header)
			sheet.write(5, 36, 'Pot Advance/Pot Koperasi', header)
			sheet.write(5, 37, 'Pot Kehadiran', header)
			sheet.write(5, 38, 'Uang Jaminan Sepatu', header)
			sheet.write(5, 39, 'Potongan Rapel', header)
			sheet.write(5, 40, 'BPJS Pensiun 1%', header)
			sheet.write(5, 41, 'Uang Jaminan Training', header)
			sheet.write(5, 42, 'Uang Jaminan Seragam', header)
			sheet.write(5, 43, 'Pot Uang Makan/Pot Medical', header)
			sheet.write(5, 44, 'Pot Asrama', header)
			sheet.write(5, 45, 'Uang Jaminan Training Sec. Guard', header)
			sheet.write(5, 46, 'Total Deduction', header)
			sheet.write(5, 47, 'Tax', header)
			sheet.write(5, 48, 'Net Salary', header)

			current_row = 6
			area_ids = payslip_ids.mapped('area_id')
			for area in area_ids:
				payslip_area = payslip_ids.filtered(lambda x: x.area_id == area)
				grade_ids = payslip_area.mapped('gerading')
				sheet.write(current_row, 0, area.name, cell_left)
				sheet.write(current_row, 1, area.code_site_card, cell_left)
				for grade in grade_ids:
					# Grade
					payslip_grade = payslip_area.filtered(lambda x: x.gerading == grade)
					man_power = len(payslip_grade.mapped('employee_id'))	
					sheet.write(current_row, 2, grade.name, cell_left)
					sheet.write(current_row, 3, man_power, cell_left)

					# Basic Salary
					salary_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'NBASICSALARY')
					total_salary = sum(salary_lines.mapped('total'))
					sheet.write(current_row, 4, total_salary, cell_left)

					# JHT 3.7%
					jht_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['ALJHT', 'ALJHT_M'])
					total_jht = sum(jht_lines.mapped('total'))
					sheet.write(current_row, 5, total_jht, cell_left)

					# JKK 0.24%
					jkk_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['ALJKK', 'ALJKK_M'])
					total_jkk = sum(jkk_lines.mapped('total'))
					sheet.write(current_row, 6, total_jkk, cell_left)

					# JKM 0.3%
					jkm_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['ALJKM', 'ALJKM_M'])
					total_jkm = sum(jkm_lines.mapped('total'))
					sheet.write(current_row, 7, total_jkm, cell_left)

					# Meal Allowance
					meal_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALMEAL')
					total_meal_allowance = sum(meal_allowance_lines.mapped('total'))
					sheet.write(current_row, 8, total_meal_allowance, cell_left)

					# Compentency Allowance
					compentency_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALCOMP')
					total_compentency_allowance = sum(compentency_allowance_lines.mapped('total'))
					sheet.write(current_row, 9, total_compentency_allowance, cell_left)

					# BPJS Pensiun 2% Perusahaan
					bpjs_pensiun_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['ALBPJSPENSION', 'ALBPJSPENSION_M'])
					total_bpjs_pensiun = sum(bpjs_pensiun_lines.mapped('total'))
					sheet.write(current_row, 10, total_bpjs_pensiun, cell_left)

					# BPJS Kesehatan 4% Perusahaan
					bpjs_kesehatan_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['ALBPJSKESCOMP', 'ALBPJSKESCOMP_M'])
					total_bpjs_kesehatan = sum(bpjs_kesehatan_lines.mapped('total'))
					sheet.write(current_row, 11, total_bpjs_kesehatan, cell_left)

					# Long Shift
					long_shift_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALLONG')
					total_long_shift = sum(long_shift_lines.mapped('total'))
					sheet.write(current_row, 12, total_long_shift, cell_left)

					# Adjustment
					adjustment_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALADJ')
					total_adjustment = sum(adjustment_lines.mapped('total'))
					sheet.write(current_row, 13, total_adjustment, cell_left)

					# Guaranted Incentive
					guaranted_incentive_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALGUARAN')
					total_guaranted_incentive = sum(guaranted_incentive_lines.mapped('total'))
					sheet.write(current_row, 14, total_guaranted_incentive, cell_left)

					# Incentive
					incentive_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALINS')
					total_incentive = sum(incentive_lines.mapped('total'))
					sheet.write(current_row, 15, total_incentive, cell_left)

					# Location Allowance
					location_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALLOC')
					total_location_allowance = sum(location_allowance_lines.mapped('total'))
					sheet.write(current_row, 16, total_location_allowance, cell_left)

					# Position Allowance
					position_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALPOS')
					total_position_allowance = sum(position_allowance_lines.mapped('total'))
					sheet.write(current_row, 17, total_position_allowance, cell_left)

					# Transport Allowance
					transport_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALTRANS')
					total_transport_allowance = sum(transport_allowance_lines.mapped('total'))
					sheet.write(current_row, 18, total_transport_allowance, cell_left)

					# Overtime
					overtime_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['OVTINDEX', 'ALOVT'])
					total_overtime = sum(overtime_lines.mapped('total'))
					sheet.write(current_row, 19, total_overtime, cell_left)

					# ALL BACKUP
					all_backup_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALBKUP')
					total_all_backup = sum(all_backup_lines.mapped('total'))
					sheet.write(current_row, 20, total_all_backup, cell_left)

					# Ins. Libur Nasional
					ins_libur_nasional_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'Alinslin')
					total_ins_libur_nasional = sum(ins_libur_nasional_lines.mapped('total'))
					sheet.write(current_row, 21, total_ins_libur_nasional, cell_left)

					# Daily Allowance
					# daily_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALDAILY')
					# total_daily_allowance = sum(daily_allowance_lines.mapped('total'))
					sheet.write(current_row, 22, 0, cell_left)

					# Housing Allowance
					housing_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALHOME')
					total_housing_allowance = sum(housing_allowance_lines.mapped('total'))
					sheet.write(current_row, 23, total_housing_allowance, cell_left)

					# Handphone Allowance
					handphone_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALHP')
					total_handphone_allowance = sum(handphone_allowance_lines.mapped('total'))
					sheet.write(current_row, 24, total_handphone_allowance, cell_left)

					# Performance Allowance
					performance_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALPERFORM')
					total_performance_allowance = sum(performance_allowance_lines.mapped('total'))
					sheet.write(current_row, 25, total_performance_allowance, cell_left)

					# Travel Allowance
					travel_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALTRAV')
					total_travel_allowance = sum(travel_allowance_lines.mapped('total'))
					sheet.write(current_row, 26, total_travel_allowance, cell_left)

					# Medical
					medical_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'MEDREIM')
					total_medical = sum(medical_lines.mapped('total'))
					sheet.write(current_row, 27, total_medical, cell_left)

					# Acting Allowance
					acting_allowance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'ALACT')
					total_acting_allowance = sum(acting_allowance_lines.mapped('total'))
					sheet.write(current_row, 28, total_acting_allowance, cell_left)

					# Total Income
					total_income_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'TOTALEARNING')
					total_income = sum(total_income_lines.mapped('total'))
					sheet.write(current_row, 29, total_income, cell_left)

					# Pot JHT 3.7%
					pot_jht_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEJHT', 'DEJHT_M'])
					total_pot_jht = sum(pot_jht_lines.mapped('total'))
					sheet.write(current_row, 30, total_pot_jht, cell_left)

					# JHT employee 2%
					jht_employee_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEJHTEMP', 'DEJHTEMP_M'])
					total_jht_employee = sum(jht_employee_lines.mapped('total'))
					sheet.write(current_row, 31, total_jht_employee, cell_left)

					# Pot JKK 0.24%
					pot_jkk_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEJKK', 'DEJKK_M'])
					total_pot_jkk = sum(pot_jkk_lines.mapped('total'))
					sheet.write(current_row, 32, total_pot_jkk, cell_left)

					# Pot JKM 0.3%
					pot_jkm_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEJKM', 'DEJKM_M'])
					total_pot_jkm = sum(pot_jkm_lines.mapped('total'))
					sheet.write(current_row, 33, total_pot_jkm, cell_left)

					# BPJS Pensiun 2% Perusahaan2
					bpjs_pensiun_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEBPJSPENSION', 'DEBPJSPENSION_M'])
					total_bpjs_pensiun = sum(bpjs_pensiun_lines.mapped('total'))
					sheet.write(current_row, 34, total_bpjs_pensiun, cell_left)

					# BPJS Kesehatan 4% Perusahaan2
					bpjs_kesehatan_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEBPJSKESCOMP', 'DEBPJSKESCOMP_M'])
					total_bpjs_kesehatan = sum(bpjs_kesehatan_lines.mapped('total'))
					sheet.write(current_row, 35, total_bpjs_kesehatan, cell_left)

					# Pot BPJS Kes 1% Employee
					pot_bpjs_kes_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEBPJSKS', 'DEBPJSKS_M'])
					total_pot_bpjs_kes = sum(pot_bpjs_kes_lines.mapped('total'))
					sheet.write(current_row, 36, total_pot_bpjs_kes, cell_left)

					# Pot Advance/Pot Koperasi
					pot_advance_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DEADV')
					total_pot_advance = sum(pot_advance_lines.mapped('total'))
					sheet.write(current_row, 37, total_pot_advance, cell_left)

					# Pot Kehadiran
					pot_kehadiran_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DEHDR')
					total_pot_kehadiran = sum(pot_kehadiran_lines.mapped('total'))
					sheet.write(current_row, 38, total_pot_kehadiran, cell_left)

					# Uang Jaminan Sepatu
					uang_jaminan_sepatu_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DESPT')
					total_uang_jaminan_sepatu = sum(uang_jaminan_sepatu_lines.mapped('total'))
					sheet.write(current_row, 39, total_uang_jaminan_sepatu, cell_left)

					# Potongan Rapel
					potongan_rapel_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DERAPEL')
					total_potongan_rapel = sum(potongan_rapel_lines.mapped('total'))
					sheet.write(current_row, 40, total_potongan_rapel, cell_left)

					# BPJS Pensiun 1%
					bpjs_pensiun_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DEBPJSPENSIONEMP', 'DEBPJSPENSIONEMP_M'])
					total_bpjs_pensiun = sum(bpjs_pensiun_lines.mapped('total'))
					sheet.write(current_row, 41, total_bpjs_pensiun, cell_left)

					# Uang Jaminan Training
					uang_jaminan_training_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DETRAIN')
					total_uang_jaminan_training = sum(uang_jaminan_training_lines.mapped('total'))
					sheet.write(current_row, 42, total_uang_jaminan_training, cell_left)

					# Uang Jaminan Seragam
					uang_jaminan_seragam_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code in ['DESRGM', 'DESERAGAM'])
					total_uang_jaminan_seragam = sum(uang_jaminan_seragam_lines.mapped('total'))
					sheet.write(current_row, 43, total_uang_jaminan_seragam, cell_left)	

					# Pot Uang Makan/Pot Medical
					pot_uang_makan_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DEMEAL')
					total_pot_uang_makan = sum(pot_uang_makan_lines.mapped('total'))
					sheet.write(current_row, 44, total_pot_uang_makan, cell_left)

					# Pot Asrama
					pot_asrama_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DEDORMIT')
					total_pot_asrama = sum(pot_asrama_lines.mapped('total'))
					sheet.write(current_row, 45, total_pot_asrama, cell_left)

					# Uang Jaminan Training Sec. Guard
					uang_jaminan_training_sec_guard_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'DETRAIN2')
					total_uang_jaminan_training_sec_guard = sum(uang_jaminan_training_sec_guard_lines.mapped('total'))
					sheet.write(current_row, 46, total_uang_jaminan_training_sec_guard, cell_left)

					# Total Deduction
					total_deduction_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'TOTALDEDUCTION')
					total_deduction = sum(total_deduction_lines.mapped('total'))
					sheet.write(current_row, 47, total_deduction, cell_left)

					# Tax
					tax_lines = payslip_grade.mapped('details_by_salary_rule_category').filtered(lambda x: x.category_id.code == 'TAX')
					total_tax = sum(tax_lines.mapped('total'))
					sheet.write(current_row, 48, total_tax, cell_left)

			#DETAIL DATA EMPLOYEE
			sheet2 						= workbook.add_worksheet(name=f"{divisi.name}")

			#Header
			sheet2.set_column(0, 1, 5)
			sheet2.set_column(1, 20, 20)
			sheet2.write(0, 0, 'No.', header)
			sheet2.write(0, 1, 'Name', header)
			sheet2.write(0, 2, 'Employee No', header)
			sheet2.write(0, 3, 'Gov Tax File No', header)
			sheet2.write(0, 4, 'Position', header)
			sheet2.write(0, 5, 'Grading', header)
			sheet2.write(0, 6, 'Department', header)
			sheet2.write(0, 7, 'Cost Center', header)
			sheet2.write(0, 8, 'cc', header)
			sheet2.write(0, 9, 'Grade', header)
			sheet2.write(0, 10, 'Employment Status', header)
			sheet2.write(0, 11, 'Join Date', header)
			sheet2.write(0, 12, 'Termination Date', header)
			sheet2.write(0, 13, 'Length of Service', header)
			sheet2.write(0, 14, 'Tax Status', header)
			sheet2.write(0, 15, 'Salary', header)
			sheet2.write(0, 16, 'JHT 3.7%', header)
			sheet2.write(0, 17, 'JKK 0.24%', header)
			sheet2.write(0, 18, 'JKM 0.3%', header)
			sheet2.write(0, 19, 'Meal Allowance', header)
			sheet2.write(0, 20, 'Compentency Allowance', header)
			sheet2.write(0, 21, 'BPJS Pensiun 2% Perusahaan', header)
			sheet2.write(0, 22, 'BPJS Kesehatan 4% Perusahaan', header)
			sheet2.write(0, 23, 'Long Shift', header)
			sheet2.write(0, 24, 'Adjustment', header)
			sheet2.write(0, 25, 'Guaranted Incentive', header)
			sheet2.write(0, 26, 'Other Allowance', header)
			sheet2.write(0, 27, 'Incentive', header)
			sheet2.write(0, 28, 'Location Allowance', header)
			sheet2.write(0, 29, 'Position Allowance', header)
			sheet2.write(0, 30, 'Transport Allowance', header)
			sheet2.write(0, 31, 'Overtime', header)
			sheet2.write(0, 32, 'ALL BACKUP', header)
			sheet2.write(0, 33, 'Ins. Libur Nasional', header)
			sheet2.write(0, 34, 'Daily Allowance', header)
			sheet2.write(0, 35, 'Housing Allowance', header)
			sheet2.write(0, 36, 'Handphone Allowance', header)
			sheet2.write(0, 37, 'Performance Allowance', header)
			sheet2.write(0, 38, 'Travel Allowance', header)
			sheet2.write(0, 39, 'Medical', header)
			sheet2.write(0, 40, 'Acting Allowance', header)
			sheet2.write(0, 41, 'Total Income', header)
			sheet2.write(0, 42, 'Pot JHT 3.7%', header)
			sheet2.write(0, 43, 'JHT employee 2%', header)
			sheet2.write(0, 44, 'Pot JKK 0.24%', header)
			sheet2.write(0, 45, 'Pot JKM 0.3%', header)
			sheet2.write(0, 46, 'BPJS Pensiun 2% Perusahaan2', header)
			sheet2.write(0, 47, 'BPJS Kesehatan 4% Perusahaan2', header)
			sheet2.write(0, 48, 'Pot BPJS Kes 1% Employee', header)
			sheet2.write(0, 49, 'Pot Advance/Pot Koperasi', header)
			sheet2.write(0, 50, 'Pot Kehadiran', header)
			sheet2.write(0, 51, 'Uang Jaminan Sepatu', header)
			sheet2.write(0, 52, 'Potongan Rapel', header)
			sheet2.write(0, 53, 'BPJS Pensiun 1%', header)
			sheet2.write(0, 54, 'Uang Jaminan Training', header)
			sheet2.write(0, 55, 'Uang Jaminan Seragam', header)
			sheet2.write(0, 56, 'Pot Uang Makan/Pot Medical', header)
			sheet2.write(0, 57, 'Pot Asrama', header)
			sheet2.write(0, 58, 'Uang Jaminan Training Sec. Guard', header)
			sheet2.write(0, 59, 'Total Deduction', header)
			sheet2.write(0, 60, 'Tax', header)
			sheet2.write(0, 61, 'Net Salary', header)
			sheet2.write(0, 62, 'NE_JHT', header)
			sheet2.write(0, 63, 'NE_BPJS_KS', header)
			sheet2.write(0, 64, 'NE_BPJS_PENSIONWAGE', header)
			sheet2.write(0, 65, 'NE_KLSRAWAT', header)
			sheet2.write(0, 65, 'Prorate New Join', header)
			sheet2.write(0, 67, 'Pay Freq', header)
			sheet2.write(0, 68, 'Hari Kerja', header)
			sheet2.write(0, 69, 'Total Overtime', header)
			sheet2.write(0, 70, 'NBASICSALARY', header)

		workbook.close()
		output.seek(0)
		file_base64 = base64.b64encode(output.read())
		output.close()
		self.excel_file = file_base64
		return {
            'type': 'ir.actions.act_url',
            'url': 'web/content/?model=hr.journal.detail.gaji.wizard&field=excel_file&download=true&id=%s&filename=%s' % (self.id, "Journal_Detail_Gaji.xlsx"),
            'target': 'self'
        }